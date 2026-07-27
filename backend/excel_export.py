"""
Phase 2: Excel Architecture & Formulas
Phase 3: Native Excel Visualizations
----------------------------------------
Builds the multi-tab workbook:
  - Monthly_Summary          (formulas referencing each tenant sheet)
  - <Tenant> (one per tenant, continuous dates + a native line chart)
  - All_Tenants_Weekends     (pastel green highlighted rows)
  - All_Tenants_Weekdays
"""
import re

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .data_parser import TenantData

HEADER_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WEEKEND_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # pastel green
DATE_FMT = "yyyy-mm-dd"
CURRENCY_FMT = '#,##0;(#,##0);"-"'

INVALID_SHEETNAME_CHARS = re.compile(r"[\\/*\[\]:?]")


def _safe_sheet_name(name: str, used: set[str]) -> str:
    clean = INVALID_SHEETNAME_CHARS.sub("", name).strip() or "Tenant"
    clean = clean[:31]
    base = clean
    i = 2
    while clean in used:
        suffix = f" ({i})"
        clean = base[: 31 - len(suffix)] + suffix
        i += 1
    used.add(clean)
    return clean


def _style_header_row(ws, row=1, ncols=4):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws, widths: dict[str, int]):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _write_tenant_sheet(wb: Workbook, tenant: TenantData, sheet_name: str):
    ws = wb.create_sheet(sheet_name)
    ws.append(["Date", "Day Name", "Day Type", "Sales"])
    _style_header_row(ws)

    for _, row in tenant.df.iterrows():
        ws.append([row["Date"], row["Day Name"], row["Day Type"], row["Sales"]])

    last_row = ws.max_row
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=1).number_format = DATE_FMT
        ws.cell(row=r, column=4).number_format = CURRENCY_FMT

    _autofit(ws, {"A": 12, "B": 12, "C": 12, "D": 16})

    # Phase 3: native line chart, Sales (Y) vs Date (X), continuous range.
    chart = LineChart()
    chart.title = f"{tenant.tenant_name} — Daily Sales"
    chart.style = 12
    chart.y_axis.title = "Sales"
    chart.x_axis.title = "Date"
    chart.x_axis.number_format = DATE_FMT
    chart.x_axis.majorTimeUnit = "days"
    chart.height = 9
    chart.width = 20

    data = Reference(ws, min_col=4, min_row=1, max_row=last_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=last_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].smooth = False

    ws.add_chart(chart, "F2")

    return ws, last_row


def _write_filtered_sheet(wb: Workbook, sheet_name: str, tenants: list[TenantData], day_type: str, highlight: bool):
    ws = wb.create_sheet(sheet_name)
    ws.append(["Tenant", "Date", "Day Name", "Day Type", "Sales"])
    _style_header_row(ws, ncols=5)

    for tenant in tenants:
        subset = tenant.df[tenant.df["Day Type"] == day_type]
        for _, row in subset.iterrows():
            ws.append([tenant.tenant_name, row["Date"], row["Day Name"], row["Day Type"], row["Sales"]])

    last_row = ws.max_row
    for r in range(2, last_row + 1):
        ws.cell(row=r, column=2).number_format = DATE_FMT
        ws.cell(row=r, column=5).number_format = CURRENCY_FMT
        if highlight:
            for c in range(1, 6):
                ws.cell(row=r, column=c).fill = WEEKEND_FILL

    _autofit(ws, {"A": 20, "B": 12, "C": 12, "D": 12, "E": 16})
    return ws


def build_workbook(tenants: list[TenantData]) -> Workbook:
    if not tenants:
        raise ValueError("No tenant data to export.")

    wb = Workbook()
    wb.remove(wb.active)  # drop default sheet; we'll add ours in order

    # --- Monthly_Summary (written first, filled with formulas once we know each tenant's sheet/row range) ---
    summary_ws = wb.create_sheet("Monthly_Summary")
    summary_ws.append(["Tenant Name", "Total Monthly Sales", "Weekday Average / Day", "Weekend/Holiday Average / Day"])
    _style_header_row(summary_ws)
    _autofit(summary_ws, {"A": 26, "B": 20, "C": 22, "D": 26})

    used_names: set[str] = set()
    tenant_sheet_info = []  # (tenant, sheet_name, last_row)

    for tenant in tenants:
        sheet_name = _safe_sheet_name(tenant.tenant_name, used_names)
        _, last_row = _write_tenant_sheet(wb, tenant, sheet_name)
        tenant_sheet_info.append((tenant, sheet_name, last_row))

    for i, (tenant, sheet_name, last_row) in enumerate(tenant_sheet_info, start=2):
        # Quote sheet name defensively (handles spaces / special chars).
        ref = f"'{sheet_name}'"
        sales_range = f"{ref}!$D$2:$D${last_row}"
        type_range = f"{ref}!$C$2:$C${last_row}"

        summary_ws.cell(row=i, column=1, value=tenant.tenant_name)
        summary_ws.cell(row=i, column=2, value=f"=SUM({sales_range})").number_format = CURRENCY_FMT
        summary_ws.cell(row=i, column=3, value=f'=AVERAGEIFS({sales_range},{type_range},"Weekday")').number_format = CURRENCY_FMT
        summary_ws.cell(row=i, column=4, value=f'=AVERAGEIFS({sales_range},{type_range},"Weekend")').number_format = CURRENCY_FMT

    # --- All_Tenants_Weekends / Weekdays ---
    _write_filtered_sheet(wb, "All_Tenants_Weekends", tenants, "Weekend", highlight=True)
    _write_filtered_sheet(wb, "All_Tenants_Weekdays", tenants, "Weekday", highlight=False)

    # Put Monthly_Summary first in tab order.
    wb.move_sheet("Monthly_Summary", offset=-len(wb.sheetnames))

    wb.active = 0
    return wb
