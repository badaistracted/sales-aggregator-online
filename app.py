# app.py
import os
import re
import uuid
import calendar as cal
from datetime import datetime, date
import pandas as pd
import pdfplumber
from pathlib import Path
from flask import Flask, request, jsonify, render_template_string, send_file
from werkzeug.utils import secure_filename
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from event_parser import try_event_parser, parse_event_file
from chart_builder import (
    chart_monthly_sales,
    chart_top_tenants,
    chart_traffic,
    chart_daily_sales,
)
from llm_writer import _build_kpis, generate_slide_text
from pptx_builder import build_pptx

app = Flask(__name__)
UPLOAD_FOLDER = Path("temp_uploads")
UPLOAD_FOLDER.mkdir(exist_ok=True)

# ══════════════════════════════════════════════════════════════
#  MONTH REFERENCE DATA
# ══════════════════════════════════════════════════════════════

MONTH_NAMES = {
    "jan": 1, "january": 1, "januari": 1,
    "feb": 2, "february": 2, "februari": 2,
    "mar": 3, "march": 3, "maret": 3,
    "apr": 4, "april": 4,
    "may": 5, "mei": 5,
    "jun": 6, "june": 6, "juni": 6,
    "jul": 7, "july": 7, "juli": 7,
    "aug": 8, "august": 8, "agustus": 8, "agu": 8, "ags": 8,
    "sep": 9, "september": 9, "sept": 9,
    "oct": 10, "october": 10, "okt": 10, "oktober": 10,
    "nov": 11, "november": 11, "nop": 11, "nopember": 11,
    "dec": 12, "december": 12, "des": 12, "desember": 12,
}

DATE_ALIASES = [
    "date", "tanggal", "tgl", "hari", "periode", "period",
    "transaction date", "trans date", "waktu", "datetime",
    "tgl transaksi", "tanggal transaksi",
]

SALES_ALIASES = [
    "sales", "penjualan", "total", "amount", "revenue",
    "total sales", "total penjualan", "jumlah", "nilai",
    "net sales", "gross sales", "omzet", "omset",
    "total amount", "sales amount", "income", "pendapatan",
]


# ══════════════════════════════════════════════════════════════
#  MONTH DETECTION (for validation — from Phase 1.6)
# ══════════════════════════════════════════════════════════════

def detect_months_in_text(text):
    text = str(text).lower().strip()
    found = set()
    if not text or text in ("nan", "none"):
        return found
    clean = text.replace(",", "").replace(".", "").strip()
    if clean.isdigit():
        return found

    for match in re.finditer(r"([a-z]{3,})[\s\-_/\.]+(\d{2,4})", text):
        name, yr = match.group(1), int(match.group(2))
        if yr < 100: yr += 2000
        mn = MONTH_NAMES.get(name)
        if mn and 2020 <= yr <= 2035:
            found.add((yr, mn))

    for match in re.finditer(r"(\d{4})[\s\-_/\.]+(\d{1,2})(?!\d)", text):
        yr, mn = int(match.group(1)), int(match.group(2))
        if 1 <= mn <= 12 and 2020 <= yr <= 2035:
            found.add((yr, mn))

    for match in re.finditer(r"(\d{4})[\s\-_/\.]+([a-z]{3,})", text):
        yr = int(match.group(1))
        mn = MONTH_NAMES.get(match.group(2))
        if mn and 2020 <= yr <= 2035:
            found.add((yr, mn))

    for match in re.finditer(r"(\d{1,2})[/\-](\d{1,2})[/\-](\d{4})", text):
        a, b, c = int(match.group(1)), int(match.group(2)), int(match.group(3))
        if 2020 <= c <= 2035:
            if 1 <= b <= 12: found.add((c, b))
            if 1 <= a <= 12 and a != b: found.add((c, a))

    for match in re.finditer(r"(\d{4})[/\-](\d{1,2})[/\-](\d{1,2})", text):
        yr, mn = int(match.group(1)), int(match.group(2))
        if 1 <= mn <= 12 and 2020 <= yr <= 2035:
            found.add((yr, mn))

    return found


def detect_months_in_excel(rows):
    found = set()
    for row in rows:
        for cell in row:
            s = str(cell).strip()
            if s and s not in ("", "nan", "None"):
                found.update(detect_months_in_text(s))
    return found


def detect_months_in_lines(lines):
    found = set()
    for line in lines:
        found.update(detect_months_in_text(line))
    return found


def validate_month(detected, target_year, target_month):
    target = (target_year, target_month)
    target_label = cal.month_name[target_month] + " " + str(target_year)

    if not detected:
        return {
            "status": "warning", "icon": "⚠️",
            "message": "No month/year detected. Cannot verify.",
            "match": False, "detected": [], "target": target_label,
        }

    detected_labels = sorted([cal.month_name[m] + " " + str(y) for y, m in detected])

    if target in detected:
        if len(detected) == 1:
            return {
                "status": "ok", "icon": "✅",
                "message": "Matches: " + target_label,
                "match": True, "detected": detected_labels, "target": target_label,
            }
        others = [l for l in detected_labels if l != target_label]
        return {
            "status": "ok_multi", "icon": "⚠️",
            "message": "Contains " + target_label + " but also has: " + ", ".join(others),
            "match": True, "detected": detected_labels, "target": target_label,
        }

    return {
        "status": "mismatch", "icon": "⚠️",
        "message": "WRONG MONTH — Expected " + target_label + " but found: " + ", ".join(detected_labels),
        "match": False, "detected": detected_labels, "target": target_label,
    }


# ══════════════════════════════════════════════════════════════
#  FILE READERS (Phase 1)
# ══════════════════════════════════════════════════════════════

def read_excel(path):
    try:
        engine = "xlrd" if path.suffix == ".xls" else "openpyxl"
        df = pd.read_excel(path, header=None, engine=engine).fillna("")
        rows = []
        for row in df.values.tolist():
            rows.append([str(c) if str(c) != "" else "" for c in row])
        return {"type": "table", "rows": rows, "cols": len(rows[0]) if rows else 0}
    except Exception as e:
        return {"type": "error", "message": "Excel Error: " + str(e)}


def smart_ocr_to_table(lines):
    """Convert OCR lines to preview table. Only merges dates together."""
    table_rows = []
    date_pattern = re.compile(
        r"(\d{1,2})\s+(January|February|March|April|May|June|July|August|"
        r"September|October|November|December|"
        r"Januari|Februari|Maret|Mei|Juni|Juli|Agustus|"
        r"Oktober|Desember)\s+(\d{4})",
        re.IGNORECASE
    )

    for line in lines:
        clean = str(line).strip()
        if not clean:
            continue
        merged = clean
        for day, month, year in date_pattern.findall(merged):
            original = day + " " + month + " " + year
            placeholder = day + "_" + month + "_" + year
            merged = merged.replace(original, placeholder, 1)

        parts = re.split(r"\s+", merged)
        restored = []
        for part in parts:
            if "_" in part and date_pattern.match(part.replace("_", " ")):
                restored.append(part.replace("_", " "))
            else:
                restored.append(part)
        table_rows.append(restored)

    if not table_rows:
        return None, 0
    max_cols = max(len(row) for row in table_rows)
    for row in table_rows:
        while len(row) < max_cols:
            row.append("")
    return table_rows, max_cols


def read_pdf_ocr(path):
    try:
        from pdf2image import convert_from_path
        import pytesseract

        images = convert_from_path(path, dpi=300)
        text_lines = []
        for img in images:
            text = pytesseract.image_to_string(img, lang="eng")
            if text:
                text_lines.extend([l.strip() for l in text.split("\n") if l.strip()])

        if text_lines:
            rows, cols = smart_ocr_to_table(text_lines)
            if rows:
                return {
                    "type": "table", "rows": rows, "cols": cols,
                    "is_ocr": True, "raw_lines": text_lines,
                }
            return {"type": "text", "lines": text_lines}

        return {"type": "error", "message": "PDF OCR: No text detected."}
    except Exception as e:
        return {"type": "error", "message": "PDF OCR Error: " + str(e)}


def read_pdf(path):
    try:
        tables_found = []
        text_lines = []

        with pdfplumber.open(path) as pdf:
            for page in pdf.pages:
                page_tables = page.extract_tables()
                if page_tables:
                    for table in page_tables:
                        for row in table:
                            cleaned = [str(c).strip() if c else "" for c in row]
                            if any(cleaned):
                                tables_found.append(cleaned)
                text = page.extract_text()
                if text:
                    text_lines.extend([l.strip() for l in text.split("\n") if l.strip()])

        if tables_found:
            max_cols = max(len(r) for r in tables_found)
            for r in tables_found:
                while len(r) < max_cols:
                    r.append("")
            return {
                "type": "table", "rows": tables_found, "cols": max_cols,
                "lines": text_lines,
            }

        if text_lines and len(text_lines) > 5:
            rows, cols = smart_ocr_to_table(text_lines)
            if rows:
                return {"type": "table", "rows": rows, "cols": cols, "lines": text_lines}
            return {"type": "text", "lines": text_lines}

        return read_pdf_ocr(path)

    except Exception:
        return read_pdf_ocr(path)


# ══════════════════════════════════════════════════════════════
#  PHASE 2 — PARSERS
# ══════════════════════════════════════════════════════════════

def parse_number(raw):
    """Parse Indonesian-style numbers: '330.685.175', '128,939,034', '5000000'."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s or s.lower() in ("nan", "none", "-", "--", "n/a"):
        return None
    s = re.sub(r"[^\d.,\-]", "", s)
    if not s or s in ("-", "--"):
        return None
    neg = bool(re.match(r"^-\d", s))
    s = s.lstrip("-")
    if not s:
        return None
    if re.fullmatch(r"\d{1,3}([.,]\d{3})+", s):
        val = float(s.replace(".", "").replace(",", ""))
    elif re.fullmatch(r"\d+", s):
        val = float(s)
    elif re.fullmatch(r"\d+[.,]\d{1,2}", s):
        val = float(s.replace(",", "."))
    else:
        digits = re.sub(r"[^\d]", "", s)
        if not digits:
            return None
        val = float(digits)
    return -val if neg else val


def parse_month_year_cell(text):
    """Parse a single cell as month-year header. Returns (year, month) or None."""
    if text is None:
        return None
    s = str(text).strip().lower()
    if not s or s in ("nan", "none"):
        return None

    m = re.fullmatch(r"([a-z]{3,})[\s\-_/\.]+(\d{2,4})", s)
    if m:
        mn = MONTH_NAMES.get(m.group(1))
        yr = int(m.group(2))
        if yr < 100: yr += 2000
        if mn and 2020 <= yr <= 2035:
            return (yr, mn)

    m = re.fullmatch(r"(\d{4})[\s\-_/\.]+(\d{1,2})", s)
    if m:
        yr, mn = int(m.group(1)), int(m.group(2))
        if 1 <= mn <= 12 and 2020 <= yr <= 2035:
            return (yr, mn)

    m = re.fullmatch(r"(\d{4})[\s\-_/\.]+([a-z]{3,})", s)
    if m:
        yr = int(m.group(1))
        mn = MONTH_NAMES.get(m.group(2))
        if mn and 2020 <= yr <= 2035:
            return (yr, mn)


    return None


def _match_alias(cell, aliases):
    c = re.sub(r"\s+", " ", str(cell).strip().lower())
    if not c or c in ("nan", "none"):
        return False
    for a in aliases:
        if c == a or c.startswith(a) or a in c:
            return True
    return False

def _norm_header(x):
    return re.sub(r"\s+", " ", str(x).strip().lower())


def _score_date_header(cell):
    """
    Strict date-header scoring.
    IMPORTANT: does NOT treat 'periode' as a date column header.
    """
    c = _norm_header(cell)

    if c in ("tgl", "tanggal", "date"):
        return 100
    if c in ("transaction date", "trans date", "tanggal transaksi", "tgl transaksi"):
        return 95
    if c.startswith("tanggal ") or c.startswith("tgl "):
        return 90
    if c in ("datetime", "waktu"):
        return 60

    return 0


def _score_sales_header(cell):
    """
    Prioritise the best sales column, not the first one found.
    Example:
      TOTAL PENJUALAN  >  NETT PENJUALAN  >  PENJUALAN
    """
    c = _norm_header(cell)

    # strongest matches
    if c in ("total penjualan", "total sales", "gross sales", "sales total", "sales amount", "total amount"):
        return 100

    # strong partial matches
    if "total penjualan" in c or "total sales" in c:
        return 98

    # next-best choices
    if c in ("nett penjualan", "net penjualan", "net sales", "nett sales", "net revenue", "gross revenue"):
        return 90
    if "nett penjualan" in c or "net sales" in c or "net revenue" in c:
        return 88

    # still useful
    if c in ("pendapatan", "revenue", "omzet", "omset"):
        return 80
    if "pendapatan" in c or "revenue" in c or "omzet" in c or "omset" in c:
        return 78

    # weakest generic matches
    if c in ("penjualan", "sales", "amount", "nilai", "jumlah"):
        return 60
    if "penjualan" in c or "sales" in c:
        return 55

    return 0


def _finance_header_bonus(row):
    """
    Bonus for rows that look like real financial header rows.
    """
    bonus = 0
    for cell in row:
        c = _norm_header(cell)
        if any(tok in c for tok in (
            "penjualan", "sales", "total", "nett", "net",
            "revenue", "pendapatan", "fnb", "phi", "tax"
        )):
            bonus += 1
    return min(bonus, 6) * 4


def _candidate_data_score(rows, header_idx, date_c, sales_c):
    """
    Validate candidate header row using the rows below it.
    We want:
    - date column to parse as dates
    - sales column to parse as numbers
    """
    checked = 0
    good_dates = 0
    good_nums = 0

    for r in rows[header_idx + 1: header_idx + 13]:
        dv = r[date_c] if date_c < len(r) else ""
        sv = r[sales_c] if sales_c < len(r) else ""

        if str(dv).strip() == "" and str(sv).strip() == "":
            continue

        checked += 1

        parsed_date = parse_date_cell(dv)
        if parsed_date is not None:
            good_dates += 1

        if parse_number(sv) is not None:
            good_nums += 1

    if checked == 0:
        return 0

    return (good_dates / checked) * 40 + (good_nums / checked) * 40


# ─── Parser 1: Excel Pivot (tenants as rows, months as columns) ──

def try_excel_pivot(rows):
    best = None
    for idx, row in enumerate(rows[:50]):
        month_cols = {}
        for c, cell in enumerate(row):
            my = parse_month_year_cell(cell)
            if my:
                month_cols[c] = my
        if len(month_cols) >= 2:
            best = (idx, month_cols)
            break

    if not best:
        return None

    header_idx, month_cols = best
    ncols = max(len(r) for r in rows) if rows else 0

    label_col = None
    for c in range(ncols):
        if c not in month_cols:
            label_col = c
            break
    if label_col is None:
        return None

    skip_kw = ("total", "grand", "jumlah", "subtotal", "sub total",
               "sum", "rata", "average", "nan", "none")
    tenants = {}

    for r in rows[header_idx + 1:]:
        if label_col >= len(r):
            continue
        name = str(r[label_col]).strip()
        if not name or name.lower() in ("nan", "none"):
            continue
        if any(k in name.lower() for k in skip_kw):
            continue
        if not re.search(r"[a-zA-Z]", name):
            continue

        monthly = {}
        for c, (yr, mn) in month_cols.items():
            if c < len(r):
                v = parse_number(r[c])
                if v is not None:
                    key = f"{yr}-{mn:02d}"
                    monthly[key] = monthly.get(key, 0) + v
        if monthly:
            tenants[name] = {"monthly": monthly, "daily": []}

    if not tenants:
        return None

    months_found = sorted({k for t in tenants.values() for k in t["monthly"]})
    return {
        "success": True,
        "format": "excel_pivot",
        "tenants": tenants,
        "message": f"Pivot: {len(tenants)} tenant(s) × {len(months_found)} month(s). "
                   f"Header at row {header_idx + 1} (title rows above ignored).",
    }


# ─── Parser 2: Excel Columnar (Date | Sales columns) ──────────

def parse_date_cell(raw):
    """
    Parse dates safely.
    - If it looks like ISO format (YYYY-MM-DD), do NOT use dayfirst=True
    - Only use dayfirst=True for ambiguous human-style dates like 01/05/2026
    """
    if raw is None:
        return None

    s = str(raw).strip()
    if not s or s.lower() in ("nan", "none", "-", "--"):
        return None

    iso_formats = [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
    ]
    for fmt in iso_formats:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    if re.match(r"^\d{4}[-/]", s):
        dt = pd.to_datetime(s, errors="coerce", yearfirst=True)
        if not pd.isna(dt):
            return dt.date()

    dt = pd.to_datetime(s, errors="coerce", dayfirst=True)
    if not pd.isna(dt):
        return dt.date()

    return None

def try_excel_columnar(rows):
    """
    Detect a classic daily table with a true header row somewhere below title/logo rows.
    Example:
      TGL | PENJUALAN | FNB | CHOO-CHOO TRAIN | TOTAL PENJUALAN | ...
    """
    best = None

    for idx, row in enumerate(rows[:50]):
        date_candidates = []
        sales_candidates = []

        for c, cell in enumerate(row):
            ds = _score_date_header(cell)
            ss = _score_sales_header(cell)

            if ds > 0:
                date_candidates.append((ds, c, str(cell)))
            if ss > 0:
                sales_candidates.append((ss, c, str(cell)))

        if not date_candidates or not sales_candidates:
            continue

        best_date_score, date_c, date_label = max(date_candidates, key=lambda x: x[0])
        best_sales_score, sales_c, sales_label = max(sales_candidates, key=lambda x: x[0])

        score = (
            best_date_score +
            best_sales_score +
            _finance_header_bonus(row) +
            _candidate_data_score(rows, idx, date_c, sales_c)
        )

        if best is None or score > best["score"]:
            best = {
                "score": score,
                "header_idx": idx,
                "date_c": date_c,
                "sales_c": sales_c,
                "date_label": date_label,
                "sales_label": sales_label,
            }

    if best is None:
        return None

    header_idx = best["header_idx"]
    date_c = best["date_c"]
    sales_c = best["sales_c"]

    daily = []

    for r in rows[header_idx + 1:]:
        if date_c >= len(r) or sales_c >= len(r):
            continue

        raw_date = str(r[date_c]).strip()
        raw_sales = r[sales_c]

        parsed_date = parse_date_cell(raw_date)
        if parsed_date is None:
            continue

        s = parse_number(raw_sales)
        if s is None:
            continue

        daily.append({
            "date": str(parsed_date),
            "sales": s,
        })

    if not daily:
        return None

    monthly = {}
    for d in daily:
        key = d["date"][:7]
        monthly[key] = monthly.get(key, 0) + d["sales"]

    return {
        "success": True,
        "format": "excel_columnar",
        "tenants": {
            "__FROM_FILENAME__": {
                "monthly": monthly,
                "daily": daily,
            }
        },
        "message": (
            f"Columnar: {len(daily)} daily rows. "
            f"Header row {header_idx + 1}. "
            f"Date='{best['date_label']}' | Sales='{best['sales_label']}'"
        ),
    }


# ─── Parser 3: PDF Daily Lines (OCR / digital text) ─────────

_MONTH_KEYS = sorted(MONTH_NAMES.keys(), key=len, reverse=True)
_MONTH_RX = "|".join(re.escape(k) for k in _MONTH_KEYS)

DAILY_RX = re.compile(
    r"^\s*\d{1,3}\s*[\.,]?\s+"           # row index: "1." or "11,"
    r"(\d{1,2})\s+"                       # day: 01
    r"(" + _MONTH_RX + r")\s+"            # month name: March
    r"(\d{4})"                            # year: 2026
    r"\s*,?\s*"
    r"([A-Za-z]+)?\s*"                    # optional day name: Sunday
    r"(.*)$",                             # rest: numbers + OCR noise
    re.IGNORECASE,
)

NUM_RX = re.compile(r"\d{1,3}(?:[.,]\d{3,})+|\d{5,}")


def try_pdf_daily(lines, filename=""):
    daily = []
    header_lines = []

    for line in lines:
        clean = str(line).strip()
        if not clean:
            continue
        m = DAILY_RX.match(clean)
        if m:
            day = int(m.group(1))
            mn = MONTH_NAMES.get(m.group(2).lower())
            yr = int(m.group(3))
            rest = m.group(5)

            if not mn or not (2020 <= yr <= 2035):
                continue
            try:
                d = date(yr, mn, day)
            except ValueError:
                continue

            nums = NUM_RX.findall(rest)
            vals = [parse_number(n) for n in nums]
            vals = [v for v in vals if v is not None]

            if vals:
                # Leftmost number = main "Total Penjualan" column
                daily.append({"date": str(d), "sales": vals[0]})
        else:
            header_lines.append(clean)

    if len(daily) < 3:
        return None

    # Tenant name: first line that isn't report metadata
    skip_kw = ("laporan", "report", "periode", "bulan", "page", "halaman",
               "untuk", "per mu", "per tanggal", "printed", "dicetak")
    tenant = None
    for hl in header_lines[:15]:
        low = hl.lower()
        if len(hl) < 3:
            continue
        if any(k in low for k in skip_kw):
            continue
        tenant = hl
        break

    if not tenant:
        tenant = Path(filename).stem if filename else "Unknown Tenant"

    monthly = {}
    for d in daily:
        key = d["date"][:7]
        monthly[key] = monthly.get(key, 0) + d["sales"]

    return {
        "success": True,
        "format": "pdf_daily",
        "tenants": {tenant: {"monthly": monthly, "daily": daily}},
        "message": f"Daily report: {len(daily)} days extracted for '{tenant}'. "
                   f"Sales = leftmost number column.",
    }

# ─── Parser 4: Traffic Data ──────────────────────────────────

INDO_DAYS = {
    "senin": "Monday", "selasa": "Tuesday", "rabu": "Wednesday",
    "kamis": "Thursday", "jumat": "Friday", "sabtu": "Saturday",
    "minggu": "Sunday",
}

INDO_MONTHS_FULL = {
    "januari": 1, "februari": 2, "maret": 3, "april": 4,
    "mei": 5, "juni": 6, "juli": 7, "agustus": 8,
    "september": 9, "oktober": 10, "november": 11, "desember": 12,
    "nopember": 11,
}

# Merge with english
ALL_MONTHS_FULL = {}
ALL_MONTHS_FULL.update(INDO_MONTHS_FULL)
ALL_MONTHS_FULL.update({
    "january": 1, "february": 2, "march": 3, "april": 4,
    "may": 5, "june": 6, "july": 7, "august": 8,
    "september": 9, "october": 10, "november": 11, "december": 12,
})


def _parse_indo_date(text):
    """
    Parse Indonesian date strings like:
      'Jumat, 19 Desember 2025'
      'Sabtu, 03 Januari 2026'
      'Kamis, 23  Januari 2026'  (extra spaces)
    Returns a date object or None.
    """
    s = str(text).strip()
    if not s:
        return None

    # Remove day name prefix: "Jumat, " or "Friday, "
    s = re.sub(r"^[A-Za-z]+,?\s*", "", s).strip()

    # Now we should have: "19 Desember 2025" or "03 Januari 2026"
    m = re.match(r"(\d{1,2})\s+([A-Za-z]+)\s+(\d{4})", s)
    if not m:
        return None

    day = int(m.group(1))
    month_name = m.group(2).lower()
    year = int(m.group(3))

    month_num = ALL_MONTHS_FULL.get(month_name)
    if not month_num:
        return None

    try:
        return date(year, month_num, day)
    except ValueError:
        return None


def _is_month_separator(row):
    """
    Detect rows like: 'DESEMBER 2025' or 'JANUARI 2026'
    These are month header separators in the traffic sheet.
    Returns (year, month) or None.
    """
    # Join all non-empty cells
    text = " ".join(str(c).strip() for c in row if str(c).strip() and str(c).strip() != "nan")
    text = text.strip()

    if not text:
        return None

    # Pattern: "DESEMBER 2025" or "JANUARI 2026"
    m = re.match(r"^([A-Za-z]+)\s+(\d{4})$", text)
    if m:
        month_name = m.group(1).lower()
        year = int(m.group(2))
        month_num = ALL_MONTHS_FULL.get(month_name)
        if month_num and 2020 <= year <= 2035:
            return (year, month_num)

    return None


def try_traffic_parser(rows):
    """
    Parse traffic Excel files with this structure:

    TRAFFIC PENGUNJUNG DARI QTY KDR MASUK
    NO | HARI/TANGGAL | MOBIL | MOTOR | BUS | HIACE | TOTAL | 140%
    DESEMBER 2025
    1  | Jumat, 19 Desember 2025 | 19,536 | 8,530 | 0 | 0 | 28,066 | 39,292
    ...
    JANUARI 2026
    1  | Kamis, 01 Januari 2026 | 16,428 | 9,056 | ...
    """

       # Step 1: Find the header row with 140% or TOTAL column
    header_idx = None
    total_col = None
    date_col = None

    for idx, row in enumerate(rows[:20]):
        found_140 = None
        found_total = None
        found_date = None

        for c, cell in enumerate(row):
            norm = _norm_header(cell)

            # Find date column
            if found_date is None and any(k in norm for k in (
                "hari", "tanggal", "tgl", "date",
                "hari/ tanggal", "hari/tanggal", "hari / tanggal",
            )):
                found_date = c

            # Find 140% column (priority)
            # Excel might store it as: "140%", "1.4", "1.40", 1.4, etc.
            raw = str(cell).strip()
            if "140" in raw or raw in ("1.4", "1.40", "1.4000", "140"):
                found_140 = c

            # Find TOTAL column (fallback)
            if norm == "total":
                found_total = c

        # If we found either 140% or TOTAL on this row, it's the header
        if found_140 is not None or found_total is not None:
            total_col = found_140 if found_140 is not None else found_total
            header_idx = idx
            if found_date is not None:
                date_col = found_date
            break

    if header_idx is None or total_col is None:
        return None

    # If no explicit date column found, assume column 1 (B)
    if date_col is None:
        date_col = 1

    # Step 2: Parse daily rows
    daily = []
    current_month_block = None

    for row in rows[header_idx + 1:]:
        # Check if this is a month separator row
        month_sep = _is_month_separator(row)
        if month_sep:
            current_month_block = month_sep
            continue

        # Check if this row has a parseable date
        if date_col >= len(row):
            continue

        raw_date = str(row[date_col]).strip()

        # Skip empty rows, summary rows
        if not raw_date or raw_date.lower() in ("nan", "none", ""):
            continue

        # Try to parse the date
        parsed_date = _parse_indo_date(raw_date)

        # Also try standard date parsing as fallback
        if parsed_date is None:
            parsed_date = parse_date_cell(raw_date)

        if parsed_date is None:
            continue

        # Get the TOTAL value
        if total_col >= len(row):
            continue

        traffic_val = parse_number(row[total_col])
        if traffic_val is None or traffic_val <= 0:
            continue

        daily.append({
            "date": str(parsed_date),
            "traffic": traffic_val,
        })

    if len(daily) < 3:
        return None

    # Step 3: Group by month
    monthly = {}
    for d in daily:
        key = d["date"][:7]
        monthly[key] = monthly.get(key, 0) + d["traffic"]

    # Get all months found
    months_found = sorted(monthly.keys())
    months_labels = ", ".join(
        f"{cal.month_abbr[int(m.split('-')[1])]}-{m.split('-')[0][2:]}"
        for m in months_found
    )

    return {
        "success": True,
        "format": "traffic",
        "daily": daily,
        "monthly": monthly,
        "message": (
            f"Traffic: {len(daily)} daily rows across {len(months_found)} month(s) "
            f"({months_labels}). "
            f"Using column '{rows[header_idx][total_col] if total_col < len(rows[header_idx]) else 'TOTAL'}'"
        ),
    }


# ─── Master Parse Dispatcher ─────────────────────────────────

def parse_report(data, ext, filename):
    rows = data.get("rows") if data.get("type") == "table" else None
    lines = data.get("lines") or data.get("raw_lines")

    if rows:
        # ── NEW: Try event parser first ──────────────────────
        # It has very specific markers (PERIODE row, location columns)
        # so it won't false-positive on sales files
        p = try_event_parser(rows, filename)
        if p:
            return p

        # Try traffic first (very specific structure)
        p = try_traffic_parser(rows)
        if p:
            return p

        # Try columnar (Date + Sales columns)
        p = try_excel_columnar(rows)
        if p:
            stem = Path(filename).stem
            if "__FROM_FILENAME__" in p["tenants"]:
                p["tenants"][stem] = p["tenants"].pop("__FROM_FILENAME__")
            return p

        # Try pivot (tenants as rows, months as columns)
        p = try_excel_pivot(rows)
        if p:
            return p

    if lines:
        p = try_pdf_daily(lines, filename)
        if p:
            return p

    if rows:
        joined = [" ".join(str(c) for c in r if str(c).strip()) for r in rows]
        p = try_pdf_daily(joined, filename)
        if p:
            return p

    return {
        "success": False,
        "format": None,
        "tenants": {},
        "message": "Could not detect report structure.",
    }


# ══════════════════════════════════════════════════════════════
#  HTML UI
# ══════════════════════════════════════════════════════════════

# HTML UI (Updated for professional consistency, integrated dropzone, and Lucide icons)
HTML_UI = r"""
<!DOCTYPE html>
<html>
<head>
    <title>Tenant Report Reader</title>
    <!-- Lucide Icons CDN -->
    <script src="https://unpkg.com/lucide@latest"></script>
    <style>
        :root {
            --bg-base: #0f172a;
            --bg-surface: #1e293b;
            --bg-subtle: #263248;
            --accent-primary: #3b82f6;
            --accent-primary-hover: #2563eb;
            --accent-secondary: #7c3aed;
            --accent-secondary-hover: #6d28d9;
            --text-main: #f1f5f9;
            --text-muted: #94a3b8;
            --border-color: #334155;
            --success: #10b981;
            --warning: #f59e0b;
            --danger: #ef4444;
        }

        body {
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, sans-serif;
            background: var(--bg-base);
            color: var(--text-main);
            padding: 40px 20px;
            margin: 0;
            line-height: 1.5;
        }
        
        .container { max-width: 1200px; margin: 0 auto; }
        
        .header-section {
            display: flex;
            align-items: center;
            gap: 12px;
            margin-bottom: 25px;
        }
        
        h1 { margin: 0; font-size: 1.85rem; font-weight: 700; color: #fff; }
        .subtitle { color: var(--text-muted); margin-top: 4px; margin-bottom: 25px; font-size: 0.95rem; }

        .config-card {
            background: var(--bg-surface);
            border: 1px solid var(--border-color);
            border-radius: 12px;
            padding: 24px;
            margin-bottom: 20px;
        }

        .config-title {
            font-size: 0.9rem;
            font-weight: 600;
            color: #93c5fd;
            text-transform: uppercase;
            letter-spacing: 0.05em;
            margin-bottom: 16px;
            display: flex;
            align-items: center;
            gap: 8px;
        }

        .config-row { display: flex; gap: 20px; align-items: center; flex-wrap: wrap; }
        .config-group { display: flex; flex-direction: column; gap: 6px; }
        .config-group label { font-size: 0.75rem; font-weight: 600; color: var(--text-muted); }
        
        .config-group select, .config-group input[type="number"] {
            background: var(--bg-subtle); 
            border: 1px solid var(--border-color); 
            border-radius: 8px;
            color: var(--text-main); 
            padding: 10px 14px; 
            font-size: 0.9rem; 
            outline: none;
            transition: border-color 0.2s;
        }
        .config-group select:focus, .config-group input[type="number"]:focus {
            border-color: var(--accent-primary);
        }
        
        .month-preview { 
            font-size: 0.9rem; 
            color: var(--text-muted); 
            background: rgba(59, 130, 246, 0.08);
            border: 1px solid rgba(59, 130, 246, 0.15);
            padding: 10px 16px;
            border-radius: 8px;
            margin-left: auto;
        }
        .month-preview b { color: var(--accent-primary); }

        /* Integrated Dropzone Container */
        .drop-zone {
            border: 2px dashed #475569; 
            border-radius: 16px; 
            padding: 48px 32px;
            text-align: center; 
            cursor: pointer;
            background: rgba(30, 41, 59, 0.4); 
            transition: all 0.2s ease-in-out;
            display: flex;
            flex-direction: column;
            align-items: center;
            gap: 12px;
        }
        .drop-zone:hover, .drop-zone.hover {
            border-color: var(--accent-primary);
            background: rgba(59, 130, 246, 0.03);
            transform: scale(1.005);
        }
        
        .drop-icon {
            width: 48px;
            height: 48px;
            color: var(--text-muted);
            transition: color 0.2s;
        }
        .drop-zone:hover .drop-icon {
            color: var(--accent-primary);
        }

        .drop-title { font-size: 1.1rem; font-weight: 600; color: #fff; margin: 0; }
        .drop-subtitle { font-size: 0.85rem; color: var(--text-muted); margin: 0; }

        .browse-row { display: flex; gap: 12px; margin-top: 10px; justify-content: center; }
        
        .browse-btn {
            display: inline-flex; 
            align-items: center; 
            gap: 8px;
            padding: 10px 18px; 
            border-radius: 8px; 
            font-size: 0.85rem;
            font-weight: 600; 
            border: none; 
            cursor: pointer; 
            transition: all 0.15s ease-in-out;
        }
        .browse-btn:active { transform: translateY(1px); }
        
        .btn-primary { background: var(--accent-primary); color: #fff; }
        .btn-primary:hover { background: var(--accent-primary-hover); }
        
        .btn-secondary { 
            background: transparent; 
            color: var(--text-muted); 
            border: 1px solid var(--border-color); 
        }
        .btn-secondary:hover { 
            background: var(--bg-subtle); 
            color: var(--text-main);
            border-color: #475569;
        }
        
        .btn-clear { 
            background: transparent; 
            color: var(--danger); 
            border: 1px solid rgba(239, 68, 68, 0.2); 
            padding: 8px 14px;
            font-size: 0.8rem;
        }
        .btn-clear:hover { 
            background: rgba(239, 68, 68, 0.08); 
            border-color: var(--danger);
        }
        
        .hidden-input { display: none; }

        .loader {
            display: none; text-align: center; color: var(--accent-primary);
            font-weight: 600; margin-top: 25px; padding: 20px;
            background: var(--bg-surface);
            border: 1px solid var(--border-color);
            border-radius: 12px;
        }
        .spinner {
            display: inline-block; width: 16px; height: 16px;
            border: 2px solid var(--accent-primary); border-top: 2px solid transparent;
            border-radius: 50%; animation: spin 0.6s linear infinite;
            margin-right: 10px; vertical-align: middle;
        }
        @keyframes spin { to { transform: rotate(360deg); } }

        /* Summary Dashboard */
        .summary { margin-top: 25px; display: none; }
        .summary-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 16px; }
        .summary-card {
            background: var(--bg-surface); 
            border: 1px solid var(--border-color);
            padding: 20px; 
            border-radius: 12px; 
            display: flex;
            align-items: center;
            gap: 16px;
        }
        .summary-card-icon {
            background: rgba(255, 255, 255, 0.03);
            border-radius: 8px;
            padding: 10px;
            display: flex;
            align-items: center;
            justify-content: center;
        }
        .summary-card-info { display: flex; flex-direction: column; }
        .summary-card .num { font-size: 1.5rem; font-weight: 700; line-height: 1.2; }
        .summary-card .label { font-size: 0.75rem; color: var(--text-muted); margin-top: 2px; text-transform: uppercase; letter-spacing: 0.05em; }
        
        .c-blue { color: var(--accent-primary); } 
        .c-green { color: var(--success); }
        .c-red { color: var(--danger); } 
        .c-yellow { color: var(--warning); }
        .c-purple { color: var(--accent-secondary); }

        /* Master Report Table */
        .master { margin-top: 25px; display: none; }
        .master-warn {
            background: rgba(245,158,11,0.06);
            border: 1px solid rgba(245,158,11,0.2);
            color: #fef08a; padding: 12px 16px; border-radius: 8px;
            font-size: 0.85rem; margin-bottom: 16px; display: none;
        }
        .master-table-wrap { overflow-x: auto; border-radius: 8px; border: 1px solid var(--border-color); }
        .master-table { width: 100%; border-collapse: collapse; font-size: 0.85rem; }
        .master-table th {
            background: #141b2d; color: #93c5fd; padding: 12px 16px;
            text-align: right; font-weight: 600; white-space: nowrap;
            border-bottom: 2px solid var(--accent-primary);
        }
        .master-table th:first-child, .master-table td:first-child { text-align: left; }
        .master-table td {
            padding: 10px 16px; border-bottom: 1px solid rgba(255,255,255,0.04);
            color: #cbd5e1; text-align: right; white-space: nowrap;
        }
        .master-table tr:hover td { background: rgba(59,130,246,0.04); }
        .master-table .t-col { background: rgba(59,130,246,0.08); }
        .master-table .t-head { background: rgba(37,99,235,0.25); color: #fff; }
        .master-table .total-row td {
            background: rgba(59,130,246,0.12); font-weight: 700;
            border-top: 2px solid var(--accent-primary);
            color: #fff;
        }
        .master-table .src { font-size: 0.75rem; color: var(--text-muted); }
        .master-table .no-data { color: #475569; }

        /* Unified File Cards */
        .file-list-header {
            margin-top: 30px;
            margin-bottom: 12px;
            display: flex;
            justify-content: space-between;
            align-items: center;
        }
        .file-list-title { font-size: 0.95rem; font-weight: 600; color: var(--text-main); display: flex; align-items: center; gap: 8px; }
        .file-list { margin-top: 10px; }
        
        .file-card {
            background: var(--bg-surface); 
            border: 1px solid var(--border-color);
            border-radius: 12px; 
            margin-bottom: 12px; 
            overflow: hidden;
            transition: border-color 0.2s;
        }
        .file-card.mismatch { border-color: var(--danger); }
        .file-card.matched { border-color: var(--success); }
        .file-card.warning { border-color: var(--warning); }
        
        .file-header {
            display: flex; justify-content: space-between; align-items: center;
            padding: 14px 20px; cursor: pointer; transition: background 0.15s;
        }
        .file-header:hover { background: var(--bg-subtle); }
        .file-info { display: flex; align-items: center; gap: 12px; flex-wrap: wrap; }

        .badge {
            font-size: 0.7rem; padding: 4px 8px; border-radius: 4px;
            font-weight: 700; letter-spacing: 0.05em; text-transform: uppercase;
        }
        .b-xlsx { background: #1e3a8a; color: #93c5fd; }
        .b-xls  { background: #581c87; color: #e9d5ff; }
        .b-pdf  { background: #78350f; color: #fde68a; }
        .b-ok   { background: rgba(16,185,129,0.1); color: var(--success); border: 1px solid rgba(16,185,129,0.25); }
        .b-fail { background: rgba(239,68,68,0.1); color: var(--danger); border: 1px solid rgba(239,68,68,0.25); }
        .b-parse { background: rgba(167,139,250,0.1); color: #c084fc; border: 1px solid rgba(167,139,250,0.25); }

        .month-badge { font-size: 0.7rem; padding: 4px 8px; border-radius: 4px; font-weight: 600; display: inline-flex; align-items: center; gap: 4px; }
        .mb-match { background: rgba(16,185,129,0.1); color: var(--success); border: 1px solid rgba(16,185,129,0.2); }
        .mb-mismatch { background: rgba(239,68,68,0.1); color: var(--danger); border: 1px solid rgba(239,68,68,0.2); }
        .mb-warn { background: rgba(245,158,11,0.1); color: var(--warning); border: 1px solid rgba(245,158,11,0.2); }

        .row-count { font-size: 0.8rem; color: var(--text-muted); }
        .arrow { color: var(--text-muted); transition: transform 0.2s; display: flex; align-items: center; }
        .arrow.open { transform: rotate(180deg); }

        .month-bar {
            padding: 10px 20px; font-size: 0.8rem;
            display: flex; align-items: center; gap: 8px;
        }
        .month-bar.match { background: rgba(16,185,129,0.04); color: #a7f3d0; border-top: 1px solid rgba(16,185,129,0.1); }
        .month-bar.mismatch { background: rgba(239,68,68,0.04); color: #fecaca; border-top: 1px solid rgba(239,68,68,0.1); }
        .month-bar.warn { background: rgba(245,158,11,0.04); color: #fef08a; border-top: 1px solid rgba(245,158,11,0.1); }
        .month-bar .detected-list { font-size: 0.85em; color: var(--text-muted); margin-left: auto; }

        .parse-bar {
            padding: 8px 20px; font-size: 0.8rem; color: #d8b4fe;
            background: rgba(167,139,250,0.03);
            border-top: 1px solid rgba(167,139,250,0.08);
            display: flex; align-items: center; gap: 6px;
        }

        .preview-area {
            display: none; border-top: 1px solid var(--border-color);
            max-height: 500px; overflow: auto; background: #0b0f19;
        }
        .data-table {
            width: 100%; border-collapse: collapse; font-size: 0.78rem;
            font-family: Menlo, Monaco, Consolas, monospace;
        }
        .data-table th {
            background: #141b2d; color: #93c5fd; padding: 8px 12px;
            text-align: left; font-weight: 600; position: sticky; top: 0;
            z-index: 1; border-bottom: 2px solid var(--accent-primary); white-space: nowrap;
        }
        .data-table td {
            padding: 6px 12px; border-bottom: 1px solid rgba(255,255,255,0.03);
            color: #cbd5e1; white-space: nowrap;
        }
        .data-table tr:hover td { background: rgba(59, 130, 246, 0.05); }
        .data-table .row-num {
            color: #475569; text-align: right; padding-right: 12px;
            font-size: 0.8em; user-select: none;
            border-right: 1px solid var(--border-color); background: #0f1322;
        }
        .data-table .cell-empty { color: #334155; font-style: italic; }

        .text-preview {
            padding: 16px; font-family: Menlo, Monaco, Consolas, monospace; font-size: 0.78rem;
            color: var(--success); white-space: pre-wrap; line-height: 1.6;
        }
        .text-preview .line-num {
            display: inline-block; width: 40px; color: #475569;
            text-align: right; margin-right: 12px; user-select: none;
        }
        .error-preview { padding: 16px; color: var(--danger); font-size: 0.85rem; display: flex; align-items: center; gap: 8px; }
                /* Warning badge */
        .b-warn { 
            background: rgba(245,158,11,0.12); 
            color: var(--warning); 
            border: 1px solid rgba(245,158,11,0.25); 
        }

        /* Parse metadata tags */
        .meta-tag {
            display: inline-flex;
            align-items: center;
            gap: 4px;
            font-size: 0.78rem;
            color: #c4b5fd;
            background: rgba(139, 92, 246, 0.06);
            border: 1px solid rgba(139, 92, 246, 0.12);
            padding: 2px 8px;
            border-radius: 4px;
            white-space: nowrap;
        }
        .meta-tag b { color: #e9d5ff; font-weight: 600; }
        .meta-sep {
            color: rgba(148, 163, 184, 0.3);
            margin: 0 2px;
            font-size: 0.9rem;
        }
        .parse-bar {
            padding: 8px 20px;
            font-size: 0.8rem;
            color: #d8b4fe;
            background: rgba(167,139,250,0.03);
            border-top: 1px solid rgba(167,139,250,0.08);
            display: flex;
            align-items: center;
            gap: 8px;
            flex-wrap: wrap;
        }
    </style>
</head>
<body>
<div class="container">
    <div class="header-section">
        <i data-lucide="building-2" style="width: 32px; height: 32px; color: var(--accent-primary);"></i>
        <div>
            <h1>Tenant Report Reader</h1>
            <div class="subtitle" style="margin:0">Automated reporting pipeline: parse uploads into unified Excel formats &amp; PPTX dashboards.</div>
        </div>
    </div>

    <!-- CONFIGURATION CARD -->
    <div class="config-card">
        <div class="config-title">
            <i data-lucide="calendar" style="width: 16px; height: 16px;"></i>
            Reporting Period Configuration
        </div>
        <div class="config-row">
            <div class="config-group">
                <label>Target Month</label>
                <select id="monthSel">
                    <option value="1">January</option><option value="2">February</option>
                    <option value="3">March</option><option value="4">April</option>
                    <option value="5">May</option><option value="6">June</option>
                    <option value="7">July</option><option value="8">August</option>
                    <option value="9">September</option><option value="10">October</option>
                    <option value="11">November</option><option value="12">December</option>
                </select>
            </div>
            <div class="config-group">
                <label>Target Year</label>
                <input id="yearIn" type="number" value="2026" min="2020" max="2100" style="width:100px"/>
            </div>
            <div class="month-preview">
                Processing reports for: <b id="targetLabel">January 2026</b>
            </div>
        </div>
    </div>

    <!-- INTEGRATED DROP ZONE -->
    <div class="drop-zone" id="dropZone">
        <i data-lucide="upload-cloud" class="drop-icon"></i>
        <div class="drop-title">Drag and drop calendar, sales, or traffic files here</div>
        <div class="drop-subtitle">Accepted types: .xlsx, .xls, .pdf · Size limit: 25MB per file</div>
        
        <div class="browse-row">
            <button class="browse-btn btn-primary" onclick="document.getElementById('fileInput').click()">
                <i data-lucide="file-plus" style="width: 14px; height: 14px;"></i>
                Browse Files
            </button>
            <button class="browse-btn btn-secondary" onclick="document.getElementById('folderInput').click()">
                <i data-lucide="folder-open" style="width: 14px; height: 14px;"></i>
                Browse Folders
            </button>
        </div>
    </div>

    <input type="file" id="fileInput" class="hidden-input" accept=".xlsx,.xls,.pdf" multiple />
    <input type="file" id="folderInput" class="hidden-input" webkitdirectory />

    <div id="loader" class="loader">
        <span class="spinner"></span> 
        Parsing inputs &amp; generating consolidated matrix summaries...
    </div>

    <!-- METRICS SUMMARY DASHBOARD -->
    <div class="summary" id="summary">
        <div class="summary-grid">
            <div class="summary-card">
                <div class="summary-card-icon"><i data-lucide="files" class="c-blue"></i></div>
                <div class="summary-card-info">
                    <div class="num c-blue" id="sTotal">0</div>
                    <div class="label">Total Uploaded</div>
                </div>
            </div>
            <div class="summary-card">
                <div class="summary-card-icon"><i data-lucide="check-circle-2" class="c-green"></i></div>
                <div class="summary-card-info">
                    <div class="num c-green" id="sOk">0</div>
                    <div class="label">Successfully Read</div>
                </div>
            </div>
            <div class="summary-card">
                <div class="summary-card-icon"><i data-lucide="alert-circle" class="c-red"></i></div>
                <div class="summary-card-info">
                    <div class="num c-red" id="sFail">0</div>
                    <div class="label">Rejected / Error</div>
                </div>
            </div>
            <div class="summary-card">
                <div class="summary-card-icon"><i data-lucide="calendar-range" class="c-purple"></i></div>
                <div class="summary-card-info">
                    <div class="num c-purple" id="sMatch">0</div>
                    <div class="label">Period Matches</div>
                </div>
            </div>
            <div class="summary-card">
                <div class="summary-card-icon"><i data-lucide="calendar-x" class="c-yellow"></i></div>
                <div class="summary-card-info">
                    <div class="num c-yellow" id="sWrong">0</div>
                    <div class="label">Mismatched Period</div>
                </div>
            </div>
        </div>
    </div>

    <!-- MASTER REPORT PREVIEW MATRIX -->
    <div class="master" id="masterSection">
        <div class="config-card">
            <div style="display:flex;justify-content:space-between;align-items:center;flex-wrap:wrap;gap:16px">
                <div class="config-title" style="margin-bottom:0">
                    <i data-lucide="layout-grid" style="width: 16px; height: 16px;"></i>
                    Preview
                </div>
                <div style="display:flex;gap:10px;flex-wrap:wrap">
                    <button class="browse-btn btn-secondary" id="exportBtn" onclick="exportExcel()">
                        <i data-lucide="download" style="width: 14px; height: 14px;"></i> Export Excel
                    </button>
                    <button class="browse-btn btn-primary" id="exportPptxBtn" onclick="exportPptx()">
                        <i data-lucide="presentation" style="width: 14px; height: 14px;"></i> Export PowerPoint
                    </button>
                </div>
            </div>
            <div class="master-warn" id="masterWarn"></div>
            <div class="master-table-wrap" style="margin-top:16px">
                <table class="master-table" id="masterTable"></table>
            </div>
        </div>
    </div>

    <!-- PROCESSED FILES QUEUE -->
    <div class="file-list-header" id="fileListHeader" style="display:none">
        <div class="file-list-title">
            <i data-lucide="layers" style="width: 16px; height: 16px; color: var(--text-muted);"></i>
            Processed Files
        </div>
        <button class="browse-btn btn-clear" onclick="clearAll()">
            <i data-lucide="rotate-ccw" style="width: 12px; height: 12px;"></i> Clear Queue
        </button>
    </div>
    <div class="file-list" id="fileList"></div>
</div>

<script>
var dropZone    = document.getElementById("dropZone");
var fileList    = document.getElementById("fileList");
var loader      = document.getElementById("loader");
var summaryEl   = document.getElementById("summary");
var fileInput   = document.getElementById("fileInput");
var folderInput = document.getElementById("folderInput");

// Initialize Lucide Icons
lucide.createIcons();

function updateLabel() {
    var months = ["","January","February","March","April","May","June",
                  "July","August","September","October","November","December"];
    var m = parseInt(document.getElementById("monthSel").value);
    var y = document.getElementById("yearIn").value;
    document.getElementById("targetLabel").textContent = months[m] + " " + y;
}
document.getElementById("monthSel").addEventListener("change", updateLabel);
document.getElementById("yearIn").addEventListener("input", updateLabel);

var now = new Date();
document.getElementById("monthSel").value = now.getMonth() + 1;
document.getElementById("yearIn").value = now.getFullYear();
updateLabel();

dropZone.addEventListener("dragover", function(e) { e.preventDefault(); dropZone.classList.add("hover"); });
dropZone.addEventListener("dragleave", function() { dropZone.classList.remove("hover"); });
dropZone.addEventListener("drop", async function(e) {
    e.preventDefault(); dropZone.classList.remove("hover");
    var items = e.dataTransfer.items;
    var formData = new FormData();
    for (var i = 0; i < items.length; i++) {
        var entry = items[i].webkitGetAsEntry();
        if (entry) await walkTree(entry, formData, "");
    }
    sendFiles(formData);
});

fileInput.addEventListener("change", function() {
    var files = fileInput.files;
    if (!files.length) return;
    var formData = new FormData();
    for (var i = 0; i < files.length; i++) {
        var name = files[i].name.toLowerCase();
        if (name.endsWith(".xlsx") || name.endsWith(".xls") || name.endsWith(".pdf"))
            formData.append("files", files[i], files[i].name);
    }
    fileInput.value = "";
    sendFiles(formData);
});

folderInput.addEventListener("change", function() {
    var files = folderInput.files;
    if (!files.length) return;
    var formData = new FormData();
    for (var i = 0; i < files.length; i++) {
        var name = files[i].name.toLowerCase();
        if (name.endsWith(".xlsx") || name.endsWith(".xls") || name.endsWith(".pdf"))
            formData.append("files", files[i], files[i].webkitRelativePath || files[i].name);
    }
    folderInput.value = "";
    sendFiles(formData);
});

function walkTree(item, formData, path) {
    return new Promise(function(resolve) {
        if (item.isFile) {
            item.file(function(file) {
                var name = file.name.toLowerCase();
                if (name.endsWith(".xlsx") || name.endsWith(".xls") || name.endsWith(".pdf"))
                    formData.append("files", file, path + file.name);
                resolve();
            });
        } else if (item.isDirectory) {
            var reader = item.createReader();
            readAllEntries(reader, function(entries) {
                var promises = [];
                for (var i = 0; i < entries.length; i++)
                    promises.push(walkTree(entries[i], formData, path + item.name + "/"));
                Promise.all(promises).then(resolve);
            });
        } else { resolve(); }
    });
}

function readAllEntries(reader, callback) {
    var all = [];
    function batch() {
        reader.readEntries(function(entries) {
            if (entries.length === 0) callback(all);
            else { all = all.concat(Array.from(entries)); batch(); }
        });
    }
    batch();
}

async function sendFiles(formData) {
    fileList.innerHTML = "";
    summaryEl.style.display = "none";
    document.getElementById("masterSection").style.display = "none";
    document.getElementById("fileListHeader").style.display = "none";
    loader.style.display = "block";

    formData.append("month", document.getElementById("monthSel").value);
    formData.append("year", document.getElementById("yearIn").value);

    try {
        var resp = await fetch("/upload", { method: "POST", body: formData });
        var data = await resp.json();
        loader.style.display = "none";
        renderAll(data);
    } catch (err) {
        loader.style.display = "none";
        alert("Error: " + err.message);
    }
}

function clearAll() {
    fileList.innerHTML = "";
    summaryEl.style.display = "none";
    document.getElementById("masterSection").style.display = "none";
    document.getElementById("fileListHeader").style.display = "none";
    loader.style.display = "none";
}

function targetKey() {
    var m = parseInt(document.getElementById("monthSel").value);
    var y = parseInt(document.getElementById("yearIn").value);
    return y + "-" + (m < 10 ? "0" + m : "" + m);
}

function renderAll(data) {
    if (data.error) { alert(data.error); return; }
    var results = data.results;
    var total = results.length, ok = 0, fail = 0, matched = 0, wrong = 0;

    for (var i = 0; i < results.length; i++) {
        var r = results[i];
        if (r.success) ok++; else fail++;
        if (r.month_check) {
            if (r.month_check.status === "mismatch") wrong++;
            else if (r.month_check.status === "ok") matched++;
        }
    }

    document.getElementById("sTotal").textContent = total;
    document.getElementById("sOk").textContent = ok;
    document.getElementById("sFail").textContent = fail;
    document.getElementById("sMatch").textContent = matched;
    document.getElementById("sWrong").textContent = wrong;
    summaryEl.style.display = "block";
    document.getElementById("fileListHeader").style.display = "flex";

    results.sort(function(a, b) {
        var order = {"mismatch": 0, "warning": 1, "ok_multi": 2, "ok": 3};
        var sa = a.month_check ? (order[a.month_check.status] || 3) : 3;
        var sb = b.month_check ? (order[b.month_check.status] || 3) : 3;
        return sa - sb;
    });

    buildMaster(results);

    for (var i = 0; i < results.length; i++) renderCard(results[i], i);
    lucide.createIcons();
}

async function exportExcel() {
    var data = window._masterExportData;
    if (!data || Object.keys(data).length === 0) {
        alert("No data to export. Upload and process files first.");
        return;
    }

    var btn = document.getElementById("exportBtn");
    var origText = btn.innerHTML;
    btn.innerHTML = '<span class="spinner"></span> Generating Excel...';
    btn.disabled = true;

    var payload = {
        month: parseInt(document.getElementById("monthSel").value),
        year: parseInt(document.getElementById("yearIn").value),
        master: data,
        traffic: window._trafficExportData || null,
        events: window._eventExportData || null
    };

    try {
        var resp = await fetch("/export", {
            method: "POST",
            headers: {"Content-Type": "application/json"},
            body: JSON.stringify(payload),
        });

        if (!resp.ok) {
            var err = await resp.json();
            alert("Export failed: " + (err.error || "Unknown error"));
            return;
        }

        var blob = await resp.blob();
        var url = URL.createObjectURL(blob);
        var a = document.createElement("a");
        a.href = url;

        var months = ["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"];
        var m = parseInt(document.getElementById("monthSel").value);
        var y = document.getElementById("yearIn").value;
        a.download = "Tenant_Report_" + months[m] + "_" + y + ".xlsx";

        document.body.appendChild(a);
        a.click();
        document.body.removeChild(a);
        URL.revokeObjectURL(url);

    } catch (err) {
        alert("Export error: " + err.message);
    } finally {
        btn.innerHTML = origText;
        btn.disabled = false;
    }
}

async function exportPptx() {
    var data = window._masterExportData;
    if (!data || Object.keys(data).length === 0) {
        alert("No data to export. Upload and process files first.");
        return;
    }

    var btn = document.getElementById("exportPptxBtn");
    var origText = btn.innerHTML;
    btn.innerHTML = '<span class="spinner"></span> Assembly Slide Deck...';
    btn.disabled = true;

    var payload = {
        month   : parseInt(document.getElementById("monthSel").value),
        year    : parseInt(document.getElementById("yearIn").value),
        master  : data,
        traffic : window._trafficExportData || null,
        events  : window._eventExportData || null
    };

    try {
        var resp = await fetch("/export_pptx", {
            method  : "POST",
            headers : {"Content-Type": "application/json"},
            body    : JSON.stringify(payload),
        });

        if (!resp.ok) {
            var err = await resp.json();
            alert("Export failed: " + (err.error || "Unknown error"));
            return;
        }

        var blob = await resp.blob();
        var url  = URL.createObjectURL(blob);
        var a    = document.createElement("a");
        a.href   = url;

        var months = ["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"];
        var m = parseInt(document.getElementById("monthSel").value);
        var y = document.getElementById("yearIn").value;
        a.download = "Mall_Report_" + months[m] + "_" + y + ".pptx";

        document.body.appendChild(a);
        a.click();
        document.body.removeChild(a);
        URL.revokeObjectURL(url);

    } catch (err) {
        alert("PowerPoint export error: " + err.message);
    } finally {
        btn.innerHTML = origText;
        btn.disabled = false;
    }
}

// ─── MASTER REPORT ─────────────────────────────────────────

function buildMaster(results) {
    var tenantMap = {};
    var allMonths = {};
    var unparsed = [];

    // Traffic data stored separately
    var trafficData = { monthly: {}, daily: [] };
    var hasTraffic = false;

    // Events data stored separately
    var eventsData = { daily: [], monthly: {}, events_flat: [] };
    var hasEvents = false;

    results.forEach(function(res) {
        if (res.parsed && res.parsed.success) {

            // Handle events files
            if (res.parsed.is_events || res.parsed.format === "events") {
                hasEvents = true;
                eventsData.daily = (eventsData.daily || []).concat(res.parsed.daily || []);
                eventsData.events_flat = (eventsData.events_flat || []).concat(res.parsed.events_flat || []);
                // Merge monthly event counts
                var em = res.parsed.monthly || {};
                Object.keys(em).forEach(function(k) {
                    if (!eventsData.monthly[k]) {
                        eventsData.monthly[k] = em[k];
                    } else {
                        eventsData.monthly[k].event_count =
                            (eventsData.monthly[k].event_count || 0) +
                            (em[k].event_count || 0);
                        eventsData.monthly[k].events =
                            (eventsData.monthly[k].events || []).concat(em[k].events || []);
                    }
                    allMonths[k] = true;
                });
                return;
            }

            // Handle traffic files
            if (res.parsed.is_traffic || res.parsed.format === "traffic") {
                hasTraffic = true;
                var tm = res.parsed.monthly || {};
                Object.keys(tm).forEach(function(k) {
                    trafficData.monthly[k] = (trafficData.monthly[k] || 0) + tm[k];
                    allMonths[k] = true;
                });
                trafficData.daily = trafficData.daily.concat(res.parsed.daily || []);
                return;
            }

            // Handle sales files
            var ts = res.parsed.tenants;
            Object.keys(ts).forEach(function(t) {
                if (!tenantMap[t]) tenantMap[t] = { monthly: {}, files: [], dailyCount: 0, daily: [] };
                if (tenantMap[t].files.indexOf(res.filename) === -1)
                    tenantMap[t].files.push(res.filename);
                var m = ts[t].monthly || {};
                Object.keys(m).forEach(function(k) {
                    if (!(k in tenantMap[t].monthly)) tenantMap[t].monthly[k] = m[k];
                    allMonths[k] = true;
                });
                var d = ts[t].daily || [];
                tenantMap[t].dailyCount += d.length;
                tenantMap[t].daily = tenantMap[t].daily.concat(d);
            });

        } else {
            var msg = res.filename;
            if (res.parsed && res.parsed.message) msg += " — " + res.parsed.message;
            unparsed.push(msg);
        }
    });

    var tenantNames = Object.keys(tenantMap);
    if (!tenantNames.length && !hasTraffic && !hasEvents) {
        document.getElementById("masterSection").style.display = "none";
        return;
    }

    var months = Object.keys(allMonths).sort();
    var tKey = targetKey();

    tenantNames.sort(function(a, b) {
        return (tenantMap[b].monthly[tKey] || 0) - (tenantMap[a].monthly[tKey] || 0);
    });

    // Build table
    var html = "<thead><tr><th>Tenant</th>";
    months.forEach(function(mk) {
        var parts = mk.split("-");
        var label = monthShort(parseInt(parts[1])) + "-" + parts[0].slice(2);
        html += '<th class="' + (mk === tKey ? "t-head" : "") + '">' + label + '</th>';
    });
    html += "<th>Total</th></tr></thead><tbody>";

    var colTotals = {};
    var grand = 0;

    // Sales rows
    tenantNames.forEach(function(t) {
        var tm = tenantMap[t];
        var rowTotal = 0;
        html += "<tr><td><b>" + esc(t) + "</b><br><span class='src'>" +
                esc(tm.files.join(", ")) +
                (tm.dailyCount ? " · " + tm.dailyCount + " daily rows" : "") +
                "</span></td>";
        months.forEach(function(mk) {
            var v = tm.monthly[mk];
            if (v !== undefined) {
                rowTotal += v;
                colTotals[mk] = (colTotals[mk] || 0) + v;
                html += '<td class="' + (mk === tKey ? "t-col" : "") + '">' + fmtNum(v) + '</td>';
            } else {
                html += '<td class="no-data ' + (mk === tKey ? "t-col" : "") + '">—</td>';
            }
        });
        grand += rowTotal;
        html += "<td><b>" + fmtNum(rowTotal) + "</b></td></tr>";
    });

    // Sales total row
    html += '<tr class="total-row"><td>💰 TOTAL SALES</td>';
    months.forEach(function(mk) {
        html += '<td class="' + (mk === tKey ? "t-col" : "") + '">' + fmtNum(colTotals[mk] || 0) + '</td>';
    });
    html += "<td>" + fmtNum(grand) + "</td></tr>";

    // Traffic row
    if (hasTraffic) {
        html += '<tr style="border-top:3px solid #3b82f6"><td><b>🚗 TRAFFIC</b><br>' +
                '<span class="src">Visitor count</span></td>';
        var trafficTotal = 0;
        months.forEach(function(mk) {
            var v = trafficData.monthly[mk];
            if (v !== undefined) {
                trafficTotal += v;
                html += '<td class="' + (mk === tKey ? "t-col" : "") + '">' + fmtNum(v) + '</td>';
            } else {
                html += '<td class="no-data ' + (mk === tKey ? "t-col" : "") + '">—</td>';
            }
        });
        html += "<td><b>" + fmtNum(trafficTotal) + "</b></td></tr>";

        // Sales per visitor row
        html += '<tr><td><b>📊 SALES / VISITOR</b><br>' +
                '<span class="src">Average spend per visitor</span></td>';
        months.forEach(function(mk) {
            var sales = colTotals[mk] || 0;
            var traffic = trafficData.monthly[mk] || 0;
            var ratio = traffic > 0 ? Math.round(sales / traffic) : 0;
            if (ratio > 0) {
                html += '<td class="' + (mk === tKey ? "t-col" : "") + '">' + fmtNum(ratio) + '</td>';
            } else {
                html += '<td class="no-data ' + (mk === tKey ? "t-col" : "") + '">—</td>';
            }
        });
        var grandRatio = (trafficTotal > 0) ? Math.round(grand / trafficTotal) : 0;
        html += "<td><b>" + fmtNum(grandRatio) + "</b></td></tr>";
    }

    // Events row
    if (hasEvents) {
        html += '<tr style="border-top:3px solid #a78bfa"><td><b>🎪 EVENTS</b><br>' +
                '<span class="src">Mall event count</span></td>';
        var totalEventCount = 0;
        months.forEach(function(mk) {
            var em = eventsData.monthly[mk];
            var count = em ? (em.event_count || 0) : 0;
            if (count > 0) {
                totalEventCount += count;
                html += '<td class="' + (mk === tKey ? "t-col" : "") + '">' + count + ' event(s)</td>';
            } else {
                html += '<td class="no-data ' + (mk === tKey ? "t-col" : "") + '">—</td>';
            }
        });
        html += "<td><b>" + totalEventCount + "</b></td></tr>";

        // Events detail row — show event names for target month only
        var targetMonthEvents = eventsData.monthly[tKey];
        if (targetMonthEvents && targetMonthEvents.events && targetMonthEvents.events.length > 0) {
            // Group event names by date for the tooltip-style detail
            var eventNames = [];
            var seen = {};
            targetMonthEvents.events.forEach(function(e) {
                var key = e.event_name;
                if (!seen[key]) {
                    seen[key] = true;
                    eventNames.push(e.event_name);
                }
            });

            html += '<tr><td colspan="' + (months.length + 2) + '" style="' +
                    'padding:8px 12px;background:rgba(167,139,250,0.05);' +
                    'font-size:0.78em;color:#a78bfa;border-top:1px solid rgba(167,139,250,0.15)">' +
                    '<b>Events this month:</b> ' +
                    esc(eventNames.slice(0, 12).join(" · ")) +
                    (eventNames.length > 12 ? ' · <i>+' + (eventNames.length - 12) + ' more</i>' : '') +
                    '</td></tr>';
        }
    }

    html += "</tbody>";

    document.getElementById("masterTable").innerHTML = html;

    var warnEl = document.getElementById("masterWarn");
    if (unparsed.length) {
        warnEl.style.display = "block";
        warnEl.innerHTML = "⚠️ <b>" + unparsed.length + " file(s) could not be parsed:</b><br>• " +
            unparsed.map(esc).join("<br>• ");
    } else {
        warnEl.style.display = "none";
    }

    document.getElementById("masterSection").style.display = "block";

    // Store for export
    window._masterExportData = tenantMap;
    window._trafficExportData = hasTraffic ? trafficData : null;
    window._eventExportData   = hasEvents  ? eventsData  : null;
}

function monthShort(m) {
    return ["","Jan","Feb","Mar","Apr","May","Jun","Jul","Aug","Sep","Oct","Nov","Dec"][m];
}

// ─── FILE CARDS RENDERER ───────────────────────────────────

function renderCard(res, idx) {
    var card = document.createElement("div");
    card.className = "file-card";

    var ext = res.filename.split(".").pop().toLowerCase();
    var extClass = ext === "pdf" ? "b-pdf" : ext === "xls" ? "b-xls" : "b-xlsx";
    var pvId = "pv_" + idx, arId = "ar_" + idx;

    // ── Determine unified status ──────────────────────────────
    var cardStatus = "neutral";  // neutral | ok | warning | error
    var statusBadgeClass = "b-ok";
    var statusText = "OK";

    if (!res.success) {
        cardStatus = "error";
        statusBadgeClass = "b-fail";
        statusText = "FAIL";
    } else if (res.month_check) {
        var ms = res.month_check.status;
        if (ms === "ok") {
            cardStatus = "ok";
            statusBadgeClass = "b-ok";
            statusText = "OK";
        } else if (ms === "mismatch") {
            cardStatus = "error";
            statusBadgeClass = "b-fail";
            statusText = "WRONG MONTH";
        } else if (ms === "ok_multi") {
            cardStatus = "warning";
            statusBadgeClass = "b-warn";
            statusText = "REVIEW";
        } else {
            cardStatus = "warning";
            statusBadgeClass = "b-warn";
            statusText = "UNVERIFIED";
        }
    }

    // Card border color
    if (cardStatus === "error") card.classList.add("mismatch");
    else if (cardStatus === "ok") card.classList.add("matched");
    else if (cardStatus === "warning") card.classList.add("warning");

    // ── Parse badge ───────────────────────────────────────────
    var parseBadge = "";
    if (res.parsed) {
        if (res.parsed.success) {
            var formatLabel = res.parsed.format;
            if (formatLabel === "events") formatLabel = "event calendar";
            else if (formatLabel === "traffic") formatLabel = "traffic data";
            else if (formatLabel === "excel_columnar") formatLabel = "daily sales";
            else if (formatLabel === "excel_pivot") formatLabel = "pivot sales";
            else if (formatLabel === "pdf_daily") formatLabel = "pdf daily";
            parseBadge = '<span class="badge b-parse"><i data-lucide="cpu" style="width:11px;height:11px;vertical-align:middle;margin-right:3px"></i>' + esc(formatLabel) + '</span>';
        } else {
            parseBadge = '<span class="badge b-fail">unparsed</span>';
        }
    }

    // ── Row count ─────────────────────────────────────────────
    var rowText = "";
    if (res.total_rows) {
        rowText = res.total_rows.toLocaleString() + " rows";
    }

    // ── Header row ────────────────────────────────────────────
    var header = document.createElement("div");
    header.className = "file-header";
    header.setAttribute("onclick", "toggle('" + pvId + "','" + arId + "')");
    header.innerHTML =
        '<div class="file-info">' +
            '<span class="badge ' + extClass + '">' + ext.toUpperCase() + '</span>' +
            '<b>' + esc(res.filename) + '</b>' +
            '<span class="badge ' + statusBadgeClass + '">' + statusText + '</span>' +
            parseBadge +
        '</div>' +
        '<div style="display:flex;align-items:center;gap:12px">' +
            (rowText ? '<span class="row-count">' + rowText + '</span>' : '') +
            '<span class="arrow" id="' + arId + '"><i data-lucide="chevron-down" style="width:16px;height:16px"></i></span>' +
        '</div>';
    card.appendChild(header);

    // ── Month validation bar (single, non-redundant) ──────────
    if (res.month_check && res.success) {
        var mc = res.month_check;
        var barClass = mc.match ? "match" : mc.status === "mismatch" ? "mismatch" : "warn";
        var iconName = mc.match && mc.status === "ok" ? "check-circle-2" : mc.status === "mismatch" ? "alert-triangle" : "alert-circle";

        // Build concise message
        var barMessage = "";
        if (mc.status === "ok") {
            barMessage = "Period verified: " + mc.target;
        } else if (mc.status === "mismatch") {
            barMessage = "Expected " + mc.target + " — file contains different period(s)";
        } else if (mc.status === "ok_multi") {
            // Summarize instead of listing every month
            var otherCount = mc.detected.length - 1;
            barMessage = "Contains " + mc.target;
            if (otherCount > 0) {
                barMessage += " + " + otherCount + " other month" + (otherCount > 1 ? "s" : "");
            }
        } else {
            barMessage = "Could not verify reporting period";
        }

        // Detected months: show compact or expandable
        var detHtml = "";
        if (mc.detected.length > 0) {
            if (mc.detected.length <= 3) {
                detHtml = '<span class="detected-list">' + mc.detected.join(", ") + '</span>';
            } else {
                var detId = "det_" + idx;
                detHtml = '<span class="detected-list">' +
                    '<span style="cursor:pointer;text-decoration:underline dotted" onclick="event.stopPropagation();var el=document.getElementById(\'' + detId + '\');el.style.display=el.style.display===\'none\'?\'inline\':\'none\'">' +
                        mc.detected.length + ' months detected — click to expand' +
                    '</span>' +
                    '<span id="' + detId + '" style="display:none;margin-left:6px">' +
                        esc(mc.detected.join(", ")) +
                    '</span>' +
                '</span>';
            }
        }

        var bar = document.createElement("div");
        bar.className = "month-bar " + barClass;
        bar.innerHTML =
            '<span style="display:flex;align-items:center;gap:6px">' +
                '<i data-lucide="' + iconName + '" style="width:13px;height:13px;flex-shrink:0"></i>' +
                esc(barMessage) +
            '</span>' +
            detHtml;
        card.appendChild(bar);
    }

    // ── Parse metadata bar (structured, not raw log) ──────────
    if (res.parsed && res.parsed.message && res.success) {
        var pm = res.parsed;
        var metaTags = [];

        // Extract structured info from the parsed result
        if (pm.format === "events" && pm.events_flat) {
            var evtMonths = {};
            pm.events_flat.forEach(function(e) { evtMonths[e.date.substring(0, 7)] = true; });
            var evtMonthKeys = Object.keys(evtMonths).sort();
            var dateRange = "";
            if (evtMonthKeys.length > 0) {
                var first = evtMonthKeys[0].split("-");
                var last = evtMonthKeys[evtMonthKeys.length - 1].split("-");
                dateRange = monthShort(parseInt(first[1])) + " " + first[0] + " – " +
                            monthShort(parseInt(last[1])) + " " + last[0];
            }
            metaTags.push('<span class="meta-tag">Events: <b>' + pm.events_flat.length + '</b></span>');
            metaTags.push('<span class="meta-tag">Months: <b>' + evtMonthKeys.length + '</b></span>');
            if (dateRange) metaTags.push('<span class="meta-tag">Range: <b>' + dateRange + '</b></span>');
        } else if (pm.format === "traffic" && pm.daily) {
            metaTags.push('<span class="meta-tag">Days: <b>' + pm.daily.length + '</b></span>');
            var trafficMonths = Object.keys(pm.monthly || {}).sort();
            metaTags.push('<span class="meta-tag">Months: <b>' + trafficMonths.length + '</b></span>');
        } else if (pm.tenants) {
            var tenantNames = Object.keys(pm.tenants);
            var totalDaily = 0;
            tenantNames.forEach(function(t) { totalDaily += (pm.tenants[t].daily || []).length; });
            if (tenantNames.length === 1) {
                metaTags.push('<span class="meta-tag">Tenant: <b>' + esc(tenantNames[0]) + '</b></span>');
            } else {
                metaTags.push('<span class="meta-tag">Tenants: <b>' + tenantNames.length + '</b></span>');
            }
            if (totalDaily > 0) metaTags.push('<span class="meta-tag">Daily rows: <b>' + totalDaily + '</b></span>');
            var salesMonths = {};
            tenantNames.forEach(function(t) {
                Object.keys(pm.tenants[t].monthly || {}).forEach(function(k) { salesMonths[k] = true; });
            });
            metaTags.push('<span class="meta-tag">Months: <b>' + Object.keys(salesMonths).length + '</b></span>');
        } else {
            // Fallback: just show the message
            metaTags.push('<span class="meta-tag">' + esc(pm.message) + '</span>');
        }

        if (metaTags.length > 0) {
            var pb = document.createElement("div");
            pb.className = "parse-bar";
            pb.innerHTML = '<i data-lucide="binary" style="width:12px;height:12px;flex-shrink:0"></i>' + metaTags.join('<span class="meta-sep">·</span>');
            card.appendChild(pb);
        }
    }

    // ── Preview area ──────────────────────────────────────────
    var preview = document.createElement("div");
    preview.className = "preview-area";
    preview.id = pvId;

    if (!res.success) {
        preview.innerHTML = '<div class="error-preview"><i data-lucide="x-octagon" style="width:16px;height:16px"></i>' + esc(res.error_message || "Unknown file read error") + '</div>';
    } else if (res.data_type === "table") {
        preview.innerHTML = buildTable(res.data_rows, res.data_cols);
    } else if (res.data_type === "text") {
        preview.innerHTML = buildText(res.data_lines);
    }

    card.appendChild(preview);
    fileList.appendChild(card);
}

function buildTable(rows, numCols) {
    var html = '<table class="data-table"><thead><tr><th class="row-num">#</th>';
    for (var c = 0; c < numCols; c++) {
        var letter;
        if (c < 26) letter = String.fromCharCode(65 + c);
        else letter = String.fromCharCode(64 + Math.floor(c/26)) + String.fromCharCode(65 + (c % 26));
        html += '<th>' + letter + '</th>';
    }
    html += '</tr></thead><tbody>';
    for (var r = 0; r < rows.length; r++) {
        html += '<tr><td class="row-num">' + (r+1) + '</td>';
        for (var c = 0; c < numCols; c++) {
            var val = c < rows[r].length ? rows[r][c] : "";
            if (val === "" || val === "nan" || val === "None")
                html += '<td class="cell-empty">&mdash;</td>';
            else
                html += '<td>' + esc(val) + '</td>';
        }
        html += '</tr>';
    }
    html += '</tbody></table>';
    return html;
}

function buildText(lines) {
    var html = '<div class="text-preview">';
    for (var i = 0; i < lines.length; i++)
        html += '<span class="line-num">' + (i+1) + '</span>' + esc(lines[i]) + '\n';
    html += '</div>';
    return html;
}

function toggle(pvId, arId) {
    var el = document.getElementById(pvId);
    var ar = document.getElementById(arId);
    if (el.style.display === "block") { el.style.display = "none"; ar.classList.remove("open"); }
    else { el.style.display = "block"; ar.classList.add("open"); }
}

function esc(text) {
    var div = document.createElement("div");
    div.textContent = String(text);
    return div.innerHTML;
}

function fmtNum(v) { return Number(v).toLocaleString("id-ID"); }
</script>
</body>
</html>
"""

# ══════════════════════════════════════════════════════════════
#  PHASE 3 — EXCEL EXPORT
# ══════════════════════════════════════════════════════════════

REPORTS_FOLDER = Path("generated_reports")
REPORTS_FOLDER.mkdir(exist_ok=True)

# Style constants
THIN_BORDER = Border(
    left=Side(style="thin", color="BDBDBD"),
    right=Side(style="thin", color="BDBDBD"),
    top=Side(style="thin", color="BDBDBD"),
    bottom=Side(style="thin", color="BDBDBD"),
)

HEADER_FILL = PatternFill("solid", fgColor="1A3A5C")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)

TITLE_FILL = PatternFill("solid", fgColor="2874A6")
TITLE_FONT = Font(color="FFFFFF", bold=True, size=14, name="Calibri")

TARGET_FILL = PatternFill("solid", fgColor="D6EAF8")
TARGET_HEADER_FILL = PatternFill("solid", fgColor="2E86C1")

TOTAL_FILL = PatternFill("solid", fgColor="EBF5FB")
TOTAL_FONT = Font(color="1B4F72", bold=True, size=11, name="Calibri")

GRAND_FILL = PatternFill("solid", fgColor="1B4F72")
GRAND_FONT = Font(color="FFFFFF", bold=True, size=11, name="Calibri")

DATA_FONT = Font(color="2C3E50", size=10, name="Calibri")
DATA_FONT_BOLD = Font(color="2C3E50", size=10, name="Calibri", bold=True)

ALT_FILL_1 = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL_2 = PatternFill("solid", fgColor="F7F9FC")

WEEKEND_FILL = PatternFill("solid", fgColor="D5F5E3")
WEEKEND_FONT = Font(color="1D6A39", size=10, name="Calibri")

NO_DATA_FONT = Font(color="BDC3C7", size=10, name="Calibri", italic=True)


def _style_cell(cell, fill=None, font=None, align=None, border=True, num_fmt=None):
    if fill:
        cell.fill = fill
    if font:
        cell.font = font
    if align:
        cell.alignment = align
    if border:
        cell.border = THIN_BORDER
    if num_fmt:
        cell.number_format = num_fmt


def _set_col_width(ws, col, width):
    ws.column_dimensions[get_column_letter(col)].width = width


def build_export_workbook(master_data, target_year, target_month,
                          traffic_data=None, events_data=None):
    """
    Build a multi-tab Excel workbook from parsed master data.

    master_data = {
        "tenant_name": {
            "monthly": {"2026-05": 123456, ...},
            "daily": [{"date": "2026-05-01", "sales": 12345}, ...],
            "files": ["file1.xlsx", ...]
        }
    }
    """
    wb = Workbook()
    wb.remove(wb.active)

    target_key = f"{target_year}-{target_month:02d}"
    month_label = f"{cal.month_name[target_month]} {target_year}"

    # Collect all months across all tenants
    all_months = sorted(set(
        k for t in master_data.values() for k in t.get("monthly", {})
    ))

    tenant_names = sorted(
        master_data.keys(),
        key=lambda t: master_data[t].get("monthly", {}).get(target_key, 0),
        reverse=True,
    )

    # ═══════════════════════════════════════════════════════════
    # Sheet 1: Monthly Summary
    # ═══════════════════════════════════════════════════════════
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_view.showGridLines = False

    # Title row
    total_cols = 1 + len(all_months) + 1
    end_col = get_column_letter(total_cols)
    ws.merge_cells(f"A1:{end_col}1")
    title_cell = ws["A1"]
    title_cell.value = f"Tenant Sales Summary — Target: {month_label}"
    _style_cell(title_cell, fill=TITLE_FILL, font=TITLE_FONT,
                align=Alignment(horizontal="center", vertical="center"))
    ws.row_dimensions[1].height = 36

    # Header row
    ws.cell(row=2, column=1, value="Tenant")
    _style_cell(ws.cell(row=2, column=1), fill=HEADER_FILL, font=HEADER_FONT, align=HEADER_ALIGN)
    _set_col_width(ws, 1, 30)

    for i, mk in enumerate(all_months):
        parts = mk.split("-")
        label = f"{cal.month_abbr[int(parts[1])]}-{parts[0][2:]}"
        col = 2 + i
        cell = ws.cell(row=2, column=col, value=label)

        if mk == target_key:
            _style_cell(cell, fill=TARGET_HEADER_FILL, font=HEADER_FONT, align=HEADER_ALIGN)
        else:
            _style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT, align=HEADER_ALIGN)

        _set_col_width(ws, col, 16)

    total_col = 2 + len(all_months)
    cell = ws.cell(row=2, column=total_col, value="Total")
    _style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT, align=HEADER_ALIGN)
    _set_col_width(ws, total_col, 18)

    # Data rows
    col_totals = {mk: 0 for mk in all_months}
    grand_total = 0

    for i, tenant in enumerate(tenant_names):
        row = 3 + i
        tm = master_data[tenant]
        monthly = tm.get("monthly", {})
        alt_fill = ALT_FILL_1 if i % 2 == 0 else ALT_FILL_2

        # Tenant name
        cell = ws.cell(row=row, column=1, value=tenant)
        _style_cell(cell, fill=alt_fill, font=DATA_FONT_BOLD,
                     align=Alignment(horizontal="left", vertical="center"))

        row_total = 0
        for j, mk in enumerate(all_months):
            col = 2 + j
            v = monthly.get(mk)

            if v is not None:
                cell = ws.cell(row=row, column=col, value=v)
                fill = TARGET_FILL if mk == target_key else alt_fill
                _style_cell(cell, fill=fill, font=DATA_FONT,
                             align=Alignment(horizontal="right"),
                             num_fmt="#,##0")
                row_total += v
                col_totals[mk] = col_totals.get(mk, 0) + v
            else:
                cell = ws.cell(row=row, column=col, value="—")
                _style_cell(cell, fill=alt_fill, font=NO_DATA_FONT,
                             align=Alignment(horizontal="center"))

        # Row total
        cell = ws.cell(row=row, column=total_col, value=row_total)
        _style_cell(cell, fill=TOTAL_FILL, font=TOTAL_FONT,
                     align=Alignment(horizontal="right"), num_fmt="#,##0")
        grand_total += row_total

    # Grand total row
    grand_row = 3 + len(tenant_names)
    cell = ws.cell(row=grand_row, column=1, value="GRAND TOTAL")
    _style_cell(cell, fill=GRAND_FILL, font=GRAND_FONT,
                 align=Alignment(horizontal="left", vertical="center"))

    for j, mk in enumerate(all_months):
        col = 2 + j
        cell = ws.cell(row=grand_row, column=col, value=col_totals.get(mk, 0))
        _style_cell(cell, fill=GRAND_FILL, font=GRAND_FONT,
                     align=Alignment(horizontal="right"), num_fmt="#,##0")

    cell = ws.cell(row=grand_row, column=total_col, value=grand_total)
    _style_cell(cell, fill=GRAND_FILL, font=GRAND_FONT,
                 align=Alignment(horizontal="right"), num_fmt="#,##0")
# Traffic row
    if traffic_data and traffic_data.get("monthly"):
        traffic_row = grand_row + 2
        ws.cell(row=grand_row + 1, column=1)  # blank separator row

        cell = ws.cell(row=traffic_row, column=1, value="🚗 TRAFFIC (Visitors)")
        _style_cell(cell, fill=PatternFill("solid", fgColor="2E86C1"),
                     font=Font(color="FFFFFF", bold=True, size=11, name="Calibri"),
                     align=Alignment(horizontal="left", vertical="center"))

        traffic_total = 0
        for j, mk in enumerate(all_months):
            col = 2 + j
            v = traffic_data["monthly"].get(mk)
            if v is not None:
                cell = ws.cell(row=traffic_row, column=col, value=v)
                fill = TARGET_FILL if mk == target_key else ALT_FILL_1
                _style_cell(cell, fill=fill, font=DATA_FONT_BOLD,
                             align=Alignment(horizontal="right"), num_fmt="#,##0")
                traffic_total += v
            else:
                cell = ws.cell(row=traffic_row, column=col, value="—")
                _style_cell(cell, fill=ALT_FILL_1, font=NO_DATA_FONT,
                             align=Alignment(horizontal="center"))

        cell = ws.cell(row=traffic_row, column=total_col, value=traffic_total)
        _style_cell(cell, fill=TOTAL_FILL, font=TOTAL_FONT,
                     align=Alignment(horizontal="right"), num_fmt="#,##0")

        # Sales per visitor row
        spv_row = traffic_row + 1
        cell = ws.cell(row=spv_row, column=1, value="📊 SALES / VISITOR")
        _style_cell(cell, fill=PatternFill("solid", fgColor="1A5276"),
                     font=Font(color="FFFFFF", bold=True, size=11, name="Calibri"),
                     align=Alignment(horizontal="left", vertical="center"))

        for j, mk in enumerate(all_months):
            col = 2 + j
            sales = col_totals.get(mk, 0)
            traffic = traffic_data["monthly"].get(mk, 0)
            ratio = round(sales / traffic) if traffic > 0 else 0

            if ratio > 0:
                cell = ws.cell(row=spv_row, column=col, value=ratio)
                fill = TARGET_FILL if mk == target_key else ALT_FILL_2
                _style_cell(cell, fill=fill, font=DATA_FONT,
                             align=Alignment(horizontal="right"), num_fmt="#,##0")
            else:
                cell = ws.cell(row=spv_row, column=col, value="—")
                _style_cell(cell, fill=ALT_FILL_2, font=NO_DATA_FONT,
                             align=Alignment(horizontal="center"))

        overall_ratio = round(grand_total / traffic_total) if traffic_total > 0 else 0
        cell = ws.cell(row=spv_row, column=total_col, value=overall_ratio)
        _style_cell(cell, fill=TOTAL_FILL, font=TOTAL_FONT,
                     align=Alignment(horizontal="right"), num_fmt="#,##0")
    ws.freeze_panes = "B3"

        # ═══════════════════════════════════════════════════════════
    # Sheet 2+: Individual Tenant Sheets (daily data)
    # ═══════════════════════════════════════════════════════════

    # Build a daily traffic lookup dict once: {"2026-05-01": 28066, ...}
    # This is mall-wide traffic, same value shown on every tenant sheet.
    daily_traffic_lookup = {}
    if traffic_data and traffic_data.get("daily"):
        for d in traffic_data["daily"]:
            date_str = d.get("date", "")
            traffic_val = d.get("traffic", 0)
            if date_str and traffic_val:
                # If multiple rows for same date, sum them
                daily_traffic_lookup[date_str] = (
                    daily_traffic_lookup.get(date_str, 0) + traffic_val
                )

    for tenant in tenant_names:
        tm = master_data[tenant]
        daily = tm.get("daily", [])

        if not daily:
            continue

        # Filter to target month only
        target_daily = [d for d in daily if d["date"].startswith(target_key)]
        if not target_daily:
            continue

        safe_name = re.sub(r"[^\w\s\-]", "", tenant)[:28]
        ws2 = wb.create_sheet(safe_name)
        ws2.sheet_view.showGridLines = False

        # Title — spans all 6 columns now
        ws2.merge_cells("A1:F1")
        title = ws2["A1"]
        title.value = f"{tenant} — Daily Sales ({month_label})"
        _style_cell(title, fill=TITLE_FILL, font=TITLE_FONT,
                    align=Alignment(horizontal="center", vertical="center"))
        ws2.row_dimensions[1].height = 32

        # Headers
        # Col 1: Date
        # Col 2: Day
        # Col 3: Day Type
        # Col 4: Sales (IDR)
        # Col 5: Mall Traffic
        # Col 6: Sales / Visitor
        headers = ["Date", "Day", "Day Type", "Sales (IDR)", "Mall Traffic", "Sales / Visitor"]
        for c, h in enumerate(headers, 1):
            cell = ws2.cell(row=2, column=c, value=h)
            _style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT, align=HEADER_ALIGN)

        _set_col_width(ws2, 1, 16)   # Date
        _set_col_width(ws2, 2, 14)   # Day
        _set_col_width(ws2, 3, 12)   # Day Type
        _set_col_width(ws2, 4, 22)   # Sales (IDR)
        _set_col_width(ws2, 5, 18)   # Mall Traffic
        _set_col_width(ws2, 6, 18)   # Sales / Visitor

        # Daily rows
        monthly_total          = 0
        monthly_traffic_total  = 0

        for i, d in enumerate(target_daily):
            row = 3 + i
            try:
                dt = datetime.strptime(d["date"], "%Y-%m-%d").date()
            except ValueError:
                continue

            day_name   = dt.strftime("%A")
            is_weekend = dt.weekday() >= 5
            sales_val  = d["sales"]

            if is_weekend:
                fill = WEEKEND_FILL
                font = WEEKEND_FONT
                day_type = "Weekend"
            else:
                fill = ALT_FILL_1 if i % 2 == 0 else ALT_FILL_2
                font = DATA_FONT
                day_type = "Weekday"

            # Col 1 — Date
            ws2.cell(row=row, column=1, value=dt)
            _style_cell(ws2.cell(row=row, column=1), fill=fill, font=font,
                        align=Alignment(horizontal="center"), num_fmt="DD-MMM-YYYY")

            # Col 2 — Day name
            ws2.cell(row=row, column=2, value=day_name)
            _style_cell(ws2.cell(row=row, column=2), fill=fill, font=font,
                        align=Alignment(horizontal="center"))

            # Col 3 — Day type
            ws2.cell(row=row, column=3, value=day_type)
            _style_cell(ws2.cell(row=row, column=3), fill=fill, font=font,
                        align=Alignment(horizontal="center"))

            # Col 4 — Sales
            ws2.cell(row=row, column=4, value=sales_val)
            _style_cell(ws2.cell(row=row, column=4), fill=fill, font=font,
                        align=Alignment(horizontal="right"), num_fmt="#,##0")

            # Col 5 — Mall Traffic
            traffic_val = daily_traffic_lookup.get(d["date"])
            if traffic_val is not None:
                ws2.cell(row=row, column=5, value=traffic_val)
                _style_cell(ws2.cell(row=row, column=5), fill=fill, font=font,
                            align=Alignment(horizontal="right"), num_fmt="#,##0")
                monthly_traffic_total += traffic_val
            else:
                # No traffic data for this date
                ws2.cell(row=row, column=5, value="—")
                _style_cell(ws2.cell(row=row, column=5), fill=fill,
                            font=NO_DATA_FONT,
                            align=Alignment(horizontal="center"))

            # Col 6 — Sales / Visitor
            if traffic_val and traffic_val > 0:
                spv = round(sales_val / traffic_val)
                ws2.cell(row=row, column=6, value=spv)
                _style_cell(ws2.cell(row=row, column=6), fill=fill, font=font,
                            align=Alignment(horizontal="right"), num_fmt="#,##0")
            else:
                ws2.cell(row=row, column=6, value="—")
                _style_cell(ws2.cell(row=row, column=6), fill=fill,
                            font=NO_DATA_FONT,
                            align=Alignment(horizontal="center"))

            monthly_total += sales_val

        # Total row
        total_row = 3 + len(target_daily)

        # Style all 6 cells in total row with grand fill first
        for c in range(1, 7):
            _style_cell(ws2.cell(row=total_row, column=c),
                        fill=GRAND_FILL, font=GRAND_FONT)

        # Col 3 — "TOTAL" label
        ws2.cell(row=total_row, column=3, value="TOTAL")
        _style_cell(ws2.cell(row=total_row, column=3), fill=GRAND_FILL, font=GRAND_FONT,
                    align=Alignment(horizontal="center"))

        # Col 4 — Total sales
        ws2.cell(row=total_row, column=4, value=monthly_total)
        _style_cell(ws2.cell(row=total_row, column=4), fill=GRAND_FILL, font=GRAND_FONT,
                    align=Alignment(horizontal="right"), num_fmt="#,##0")

        # Col 5 — Total traffic (only meaningful if traffic data exists)
        if monthly_traffic_total > 0:
            ws2.cell(row=total_row, column=5, value=monthly_traffic_total)
            _style_cell(ws2.cell(row=total_row, column=5), fill=GRAND_FILL, font=GRAND_FONT,
                        align=Alignment(horizontal="right"), num_fmt="#,##0")
        else:
            ws2.cell(row=total_row, column=5, value="—")
            _style_cell(ws2.cell(row=total_row, column=5), fill=GRAND_FILL,
                        font=Font(color="FFFFFF", bold=True, size=11, name="Calibri"),
                        align=Alignment(horizontal="center"))

        # Col 6 — Overall Sales / Visitor for the month
        if monthly_traffic_total > 0:
            overall_spv = round(monthly_total / monthly_traffic_total)
            ws2.cell(row=total_row, column=6, value=overall_spv)
            _style_cell(ws2.cell(row=total_row, column=6), fill=GRAND_FILL, font=GRAND_FONT,
                        align=Alignment(horizontal="right"), num_fmt="#,##0")
        else:
            ws2.cell(row=total_row, column=6, value="—")
            _style_cell(ws2.cell(row=total_row, column=6), fill=GRAND_FILL,
                        font=Font(color="FFFFFF", bold=True, size=11, name="Calibri"),
                        align=Alignment(horizontal="center"))

        ws2.freeze_panes = "A3"

    # ═══════════════════════════════════════════════════════════
    # Sheet: Events (if events data provided)
    # ═══════════════════════════════════════════════════════════
    if events_data and events_data.get("events_flat"):
        ws_evt = wb.create_sheet("Events")
        ws_evt.sheet_view.showGridLines = False

        # Title
        ws_evt.merge_cells("A1:D1")
        title = ws_evt["A1"]
        title.value = f"Event Calendar — {month_label}"
        _style_cell(title, fill=TITLE_FILL, font=TITLE_FONT,
                    align=Alignment(horizontal="center", vertical="center"))
        ws_evt.row_dimensions[1].height = 32

        # Headers
        evt_headers = ["Date", "Event Name", "Location", "Category"]
        for c, h in enumerate(evt_headers, 1):
            cell = ws_evt.cell(row=2, column=c, value=h)
            _style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT, align=HEADER_ALIGN)

        _set_col_width(ws_evt, 1, 16)
        _set_col_width(ws_evt, 2, 55)
        _set_col_width(ws_evt, 3, 20)
        _set_col_width(ws_evt, 4, 18)

        # Filter to target month
        target_events = [e for e in events_data["events_flat"]
                         if e["date"].startswith(target_key)]

        for i, evt in enumerate(target_events):
            row = 3 + i
            fill = ALT_FILL_1 if i % 2 == 0 else ALT_FILL_2

            try:
                dt = datetime.strptime(evt["date"], "%Y-%m-%d").date()
            except ValueError:
                dt = evt["date"]

            ws_evt.cell(row=row, column=1, value=dt)
            _style_cell(ws_evt.cell(row=row, column=1), fill=fill, font=DATA_FONT,
                        align=Alignment(horizontal="center"),
                        num_fmt="DD-MMM-YYYY" if isinstance(dt, date) else None)

            ws_evt.cell(row=row, column=2, value=evt["event_name"])
            _style_cell(ws_evt.cell(row=row, column=2), fill=fill, font=DATA_FONT,
                        align=Alignment(horizontal="left"))

            ws_evt.cell(row=row, column=3, value=evt["location"])
            _style_cell(ws_evt.cell(row=row, column=3), fill=fill, font=DATA_FONT,
                        align=Alignment(horizontal="center"))

            ws_evt.cell(row=row, column=4, value=evt.get("category", ""))
            _style_cell(ws_evt.cell(row=row, column=4), fill=fill, font=DATA_FONT,
                        align=Alignment(horizontal="center"))

        ws_evt.freeze_panes = "A3"

        # ── Add EVENTS column (G) to each tenant daily sheet ──
        # Build date -> events lookup, normalizing date format
        events_by_date = {}
        for evt in events_data.get("events_flat", []):
            d = evt["date"]
            # Normalize to YYYY-MM-DD with zero padding
            try:
                dt = datetime.strptime(d, "%Y-%m-%d").date()
                d = str(dt)  # guaranteed "2026-05-01" format
            except ValueError:
                pass
            if d not in events_by_date:
                events_by_date[d] = []
            events_by_date[d].append(evt["event_name"])

        events_lookup = {
            d: " | ".join(names)
            for d, names in events_by_date.items()
        }

        for tenant in tenant_names:
            tm = master_data[tenant]
            daily = tm.get("daily", [])
            if not daily:
                continue
            target_daily = [d for d in daily if d["date"].startswith(target_key)]
            if not target_daily:
                continue

            safe_name = re.sub(r"[^\w\s\-]", "", tenant)[:28]
            if safe_name not in [ws.title for ws in wb.worksheets]:
                continue

            ws_tenant = wb[safe_name]

            # Add EVENTS header in column G
            cell = ws_tenant.cell(row=2, column=7, value="Events")
            _style_cell(cell, fill=HEADER_FILL, font=HEADER_FONT, align=HEADER_ALIGN)
            _set_col_width(ws_tenant, 7, 45)

            # Fill events for each daily row
            # Fill events for each daily row
            for i, d in enumerate(target_daily):
                row = 3 + i
                date_str = d["date"]
                # Normalize sales date to match events lookup
                try:
                    dt_norm = datetime.strptime(date_str, "%Y-%m-%d").date()
                    date_str = str(dt_norm)
                except ValueError:
                    pass
                event_text = events_lookup.get(date_str, "-")

                try:
                    dt = datetime.strptime(date_str, "%Y-%m-%d").date()
                except ValueError:
                    dt = None

                is_weekend = dt.weekday() >= 5 if dt else False
                fit_fill = WEEKEND_FILL if is_weekend else (ALT_FILL_1 if i % 2 == 0 else ALT_FILL_2)

                ws_tenant.cell(row=row, column=7, value=event_text)
                _style_cell(
                    ws_tenant.cell(row=row, column=7),
                    fill=fit_fill,
                    font=Font(color="2C3E50", size=9, name="Calibri"),
                    align=Alignment(horizontal="left", wrap_text=True),
                )

            # Total row
            total_row = 3 + len(target_daily)
            ws_tenant.cell(row=total_row, column=7, value="—")
            _style_cell(ws_tenant.cell(row=total_row, column=7),
                        fill=GRAND_FILL, font=GRAND_FONT,
                        align=Alignment(horizontal="center"))

    # ══════════════════════════════════════════════════════════
    #  THIS MUST BE OUTSIDE ALL IF BLOCKS — always return wb
    # ══════════════════════════════════════════════════════════
    return wb

# ══════════════════════════════════════════════════════════════
#  ROUTES
# ══════════════════════════════════════════════════════════════

@app.route("/")
def index():
    return render_template_string(HTML_UI)


@app.route("/upload", methods=["POST"])
def upload():
    if "files" not in request.files:
        return jsonify({"error": "No files found"}), 400

    files = request.files.getlist("files")

    try:
        target_month = int(request.form.get("month", 1))
        target_year  = int(request.form.get("year", datetime.now().year))
    except ValueError:
        target_month = 1
        target_year  = datetime.now().year

    results = []

    for f in files:
        sid = str(uuid.uuid4())[:8]
        safe = sid + "_" + secure_filename(f.filename)
        filepath = UPLOAD_FOLDER / safe
        f.save(filepath)

        ext = filepath.suffix.lower()
        entry = {
            "filename": f.filename,
            "success": False,
            "total_rows": 0,
            "data_type": "error",
            "error_message": "",
            "month_check": None,
            "parsed": None,
        }

        if ext in [".xlsx", ".xls"]:
            # ── Try event calendar first (multi-sheet) ──────
            event_result = parse_event_file(filepath)
            if event_result and event_result.get("success"):
                # Event file detected — show confirmation text, not raw table
                event_count = len(event_result.get("events_flat", []))
                months_found = sorted(event_result.get("monthly", {}).keys())
                
                # Build a simple confirmation preview
                preview_lines = [
                    "✅ EVENT CALENDAR DETECTED",
                    "",
                    f"Total events: {event_count}",
                    f"Months covered: {', '.join(months_found)}",
                    "",
                ]
                
                # Show event count per month
                for mk in months_found:
                    mv = event_result["monthly"][mk]
                    parts = mk.split("-")
                    label = f"{cal.month_name[int(parts[1])]} {parts[0]}"
                    preview_lines.append(
                        f"  {label}: {mv['event_count']} event(s) across {mv['event_days']} day(s)"
                    )
                
                preview_lines.append("")
                preview_lines.append("─" * 50)
                preview_lines.append("")
                
                # Show sample events (first 15)
                preview_lines.append("Sample events:")
                seen_events = set()
                for evt in event_result.get("events_flat", [])[:30]:
                    name = evt["event_name"]
                    if name not in seen_events:
                        seen_events.add(name)
                        preview_lines.append(f"  📅 {evt['date']}  |  {evt['location']}  |  {name}")
                    if len(seen_events) >= 15:
                        remaining = event_count - 15
                        if remaining > 0:
                            preview_lines.append(f"  ... and {remaining} more events")
                        break
                
                entry["data_type"]  = "text"
                entry["data_lines"] = preview_lines
                entry["total_rows"] = event_count
                
                entry["success"] = True
                entry["parsed"] = event_result
                
                # Month check from events dates
                detected = set()
                for evt in event_result.get("events_flat", []):
                    try:
                        parts = evt["date"].split("-")
                        yr, mn = int(parts[0]), int(parts[1])
                        detected.add((yr, mn))
                    except (ValueError, IndexError):
                        pass
                detected.update(detect_months_in_text(f.filename))
                entry["month_check"] = validate_month(detected, target_year, target_month)
                
                results.append(entry)
                if filepath.exists():
                    os.remove(filepath)
                continue  # skip the rest of the loop for this file
            
            # ── Normal Excel reading ────────────────────────
            data = read_excel(filepath)
        elif ext == ".pdf":
            data = read_pdf(filepath)
        else:
            data = {"type": "error", "message": "Unsupported type: " + ext}

        if data["type"] == "error":
            entry["error_message"] = data.get("message", "Unknown error")
        else:
            entry["success"] = True

            if data["type"] == "table":
                entry["data_type"]  = "table"
                entry["data_rows"]  = data["rows"]
                entry["data_cols"]  = data["cols"]
                entry["total_rows"] = len(data["rows"])
                detected = detect_months_in_excel(data["rows"])
            else:
                entry["data_type"]  = "text"
                entry["data_lines"] = data["lines"]
                entry["total_rows"] = len(data["lines"])
                detected = detect_months_in_lines(data["lines"])

            # Month detection also uses raw lines if available
            if data.get("lines"):
                detected.update(detect_months_in_lines(data["lines"]))
            if data.get("raw_lines"):
                detected.update(detect_months_in_lines(data["raw_lines"]))
            detected.update(detect_months_in_text(f.filename))
            entry["month_check"] = validate_month(detected, target_year, target_month)

            # ─── PHASE 2: Parse the report structure ────────────
            parsed = parse_report(data, ext, f.filename)
            entry["parsed"] = parsed

            # If it's events data, flag and store separately
            if parsed.get("is_events"):
                entry["parsed"]["is_events"] = True
                entry["is_events_file"] = True

            # If it's traffic data, store separately
            if parsed.get("format") == "traffic":
                entry["parsed"]["is_traffic"] = True

        results.append(entry)

        if filepath.exists():
            os.remove(filepath)

    return jsonify({"results": results})

@app.route("/export", methods=["POST"])
def export():
    """Build and download the Excel report from parsed data."""
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data received"}), 400

        target_month = int(data.get("month", 1))
        target_year  = int(data.get("year", 2026))
        master_data  = data.get("master", {})
        traffic_data = data.get("traffic")
        events_data  = data.get("events")

        if not master_data and not traffic_data:
            return jsonify({"error": "No data to export"}), 400

        wb = build_export_workbook(
            master_data,
            target_year,
            target_month,
            traffic_data=traffic_data,
            events_data=events_data,
        )

        month_name = cal.month_abbr[target_month]
        filename   = f"Tenant_Report_{month_name}_{target_year}.xlsx"
        filepath   = REPORTS_FOLDER / filename

        wb.save(filepath)

        return send_file(
            filepath,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        return jsonify({"error": "Export failed: " + str(e)}), 500

@app.route("/export_pptx", methods=["POST"])
def export_pptx():
    """
    Generate and download the monthly PowerPoint.

    Expects same JSON body as /export, plus optional "openai_key".
    """
    try:
        data = request.get_json()
        if not data:
            return jsonify({"error": "No data received"}), 400

        target_month  = int(data.get("month", 1))
        target_year   = int(data.get("year", 2026))
        master_data   = data.get("master", {})
        traffic_data  = data.get("traffic")
        events_data   = data.get("events")
        target_key    = f"{target_year}-{target_month:02d}"
        month_label   = f"{cal.month_name[target_month]} {target_year}"

        if not master_data:
            return jsonify({"error": "No sales data to build report"}), 400

        # ── Step 1: Compute KPIs (pure Python) ────────────────
        kpis = _build_kpis(master_data, traffic_data, target_year, target_month, events_data)

        # ── Step 2: LLM writes commentary ─────────────────────
        # Optionally allow front-end to pass an API key
        if data.get("openai_key"):
            import os
            os.environ["OPENAI_API_KEY"] = data["openai_key"]

        llm_text = generate_slide_text(kpis)

        # ── Step 3: Build charts ───────────────────────────────
        # Monthly sales totals across all tenants
        monthly_totals = {}
        for tm in master_data.values():
            for mk, v in tm.get("monthly", {}).items():
                monthly_totals[mk] = monthly_totals.get(mk, 0) + v

        # Sales per tenant for target month
        tenant_sales_target = {
            tenant: tm.get("monthly", {}).get(target_key, 0)
            for tenant, tm in master_data.items()
            if tm.get("monthly", {}).get(target_key, 0) > 0
        }

        # All daily rows combined
        all_daily = []
        for tm in master_data.values():
            all_daily.extend(tm.get("daily", []))

        charts = {
            "monthly_sales": chart_monthly_sales(monthly_totals, target_key)
                             if monthly_totals else None,

            "top_tenants"  : chart_top_tenants(tenant_sales_target, target_key)
                             if tenant_sales_target else None,

            "traffic"      : chart_traffic(
                                 traffic_data.get("monthly", {}) if traffic_data else {},
                                 monthly_totals,
                                 target_key,
                             ) if traffic_data else None,

            "daily_sales"  : chart_daily_sales(
                                 all_daily,
                                 target_key,
                                 traffic_daily=traffic_data.get("daily", []) if traffic_data else None,
                             )
                             if all_daily else None,
        }

        # ── Step 4: Assemble PowerPoint ────────────────────────
        pptx_buf = build_pptx(
        kpis, llm_text, charts, month_label,
        target_key=target_key,
        events_data=events_data,
    )

        # ── Step 5: Return file ────────────────────────────────
        month_abbr = cal.month_abbr[target_month]
        filename   = f"Mall_Report_{month_abbr}_{target_year}.pptx"

        return send_file(
            pptx_buf,
            mimetype=(
                "application/vnd.openxmlformats-officedocument"
                ".presentationml.presentation"
            ),
            as_attachment=True,
            download_name=filename,
        )

    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({"error": f"PPTX export failed: {str(e)}"}), 500


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    print("Starting on port", port)
    app.run(host="0.0.0.0", port=port, debug=False)
